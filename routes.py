from flask import render_template, request, redirect, url_for, flash, jsonify, session, make_response, send_file
from flask_login import login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta, date
from sqlalchemy import func, desc, extract
from app import app, db
from models import (Institution, Account, Transaction, Category, Department, 
                   Vendor, MappingRule, Contract, AuditLog, Alert, Consent, User)
from utils import apply_classification_rules
import json
import re
import pandas as pd
import os
import tempfile
from werkzeug.utils import secure_filename
import io

# 언어 텍스트 사전
TEXTS = {
    'ko': {
        # 네비게이션
        'brand': 'Vlan24',
        'dashboard': '대시보드',
        'connections': '연결 관리',
        'transactions': '거래 내역',
        'rules': '분류 규칙',
        'contracts': '계약 관리',
        'departments': '부서 관리',
        'budgets': '예산 관리',
        'reports': '리포트',
        'settings': '알림/설정',
        'users': '계정 관리',
        'audit': '감사 로그',
        'sample_data': '샘플 데이터',
        'data_management': '데이터 관리',
        'logout': '로그아웃',
        'profile': '프로필',
        'admin_menu': '관리자 메뉴',
        
        # 대시보드
        'real_time_financial_status': '실시간 재무 현황 및 주요 지표',
        'current_month_income': '이번달 수입',
        'current_month_expense': '이번달 지출',
        'current_balance': '당월 수지',
        'unclassified_transactions': '미분류 거래',
        'last_month': '전월',
        'classify_now': '분류 작업 하기',
        'recent_7_days_cashflow': '최근 7일 현금흐름',
        'department_expense_status': '부서별 지출 현황',
        'recent_transactions': '최근 거래 내역',
        'recent_alerts': '최근 알림',
        'view_all': '전체 보기',
        'no_transactions': '거래 내역이 없습니다.',
        'no_alerts': '새로운 알림이 없습니다.',
        'departments_count': '부서 수',
        'total_departments': '총 부서',
        
        # 데이터 관리
        'data_import': '데이터 가져오기',
        'file_upload': '파일 업로드',
        'upload_bank_transactions': '은행거래 파일 업로드',
        'supported_formats': '지원 형식: CSV, Excel (XLS, XLSX)',
        'choose_file': '파일 선택',
        'upload': '업로드',
        'sample_download': '샘플 파일 다운로드',
        'download_csv_sample': 'CSV 샘플 다운로드',
        'download_excel_sample': 'Excel 샘플 다운로드',
        'upload_instructions': '파일 업로드 안내',
        'upload_guide_1': '1. 은행에서 거래내역을 CSV 또는 Excel 형식으로 다운로드하세요',
        'upload_guide_2': '2. 아래 샘플 파일 형식을 참고하여 데이터를 정리하세요',
        'upload_guide_3': '3. 파일을 선택하고 업로드 버튼을 클릭하세요',
        'upload_guide_4': '4. 업로드된 거래 데이터는 자동으로 분류됩니다',
        
        # 거래 관리
        'add_transaction': '거래 추가',
        'transaction_type': '거래 유형',
        'deposit': '입금',
        'withdrawal': '출금',
        'transfer': '이체',
        'amount': '금액',
        'counterparty': '거래처',
        'memo': '메모',
        'transaction_date': '거래일자',
        'add': '추가',
        'cancel': '취소',
        
        # 공통
        'language': '언어'
    },
    'en': {
        # 네비게이션
        'brand': 'Vlan24',
        'dashboard': 'Dashboard',
        'connections': 'Connections',
        'transactions': 'Transactions',
        'rules': 'Rules',
        'contracts': 'Contracts',
        'departments': 'Departments',
        'budgets': 'Budgets',
        'reports': 'Reports',
        'settings': 'Settings',
        'users': 'Users',
        'audit': 'Audit Log',
        'sample_data': 'Sample Data',
        'data_management': 'Data Management',
        'logout': 'Logout',
        'profile': 'Profile',
        'admin_menu': 'Admin Menu',
        
        # 대시보드
        'real_time_financial_status': 'Real-time Financial Status & Key Metrics',
        'current_month_income': 'Current Month Income',
        'current_month_expense': 'Current Month Expenses',
        'current_balance': 'Current Balance',
        'unclassified_transactions': 'Unclassified Transactions',
        'last_month': 'Last Month',
        'classify_now': 'Classify Now',
        'recent_7_days_cashflow': 'Recent 7 Days Cash Flow',
        'department_expense_status': 'Department Expense Status',
        'recent_transactions': 'Recent Transactions',
        'recent_alerts': 'Recent Alerts',
        'view_all': 'View All',
        'no_transactions': 'No transactions available.',
        'no_alerts': 'No new alerts.',
        'departments_count': 'Departments',
        'total_departments': 'Total Departments',
        
        # 데이터 관리
        'data_import': 'Data Import',
        'file_upload': 'File Upload',
        'upload_bank_transactions': 'Upload Bank Transaction File',
        'supported_formats': 'Supported formats: CSV, Excel (XLS, XLSX)',
        'choose_file': 'Choose File',
        'upload': 'Upload',
        'sample_download': 'Sample File Download',
        'download_csv_sample': 'Download CSV Sample',
        'download_excel_sample': 'Download Excel Sample',
        'upload_instructions': 'File Upload Instructions',
        'upload_guide_1': '1. Download transaction history from your bank in CSV or Excel format',
        'upload_guide_2': '2. Organize your data according to the sample file format below',
        'upload_guide_3': '3. Select your file and click the upload button',
        'upload_guide_4': '4. Uploaded transaction data will be automatically classified',
        
        # 거래 관리
        'add_transaction': 'Add Transaction',
        'transaction_type': 'Transaction Type',
        'deposit': 'Deposit',
        'withdrawal': 'Withdrawal',
        'transfer': 'Transfer',
        'amount': 'Amount',
        'counterparty': 'Counterparty',
        'memo': 'Memo',
        'transaction_date': 'Transaction Date',
        'add': 'Add',
        'cancel': 'Cancel',
        
        # 공통
        'language': 'Language'
    }
}

def get_text(key, lang=None):
    """언어별 텍스트 반환"""
    if lang is None:
        lang = session.get('language', 'ko')
    return TEXTS.get(lang, TEXTS['ko']).get(key, key)

# 템플릿에서 사용할 수 있도록 컨텍스트 프로세서 등록
@app.context_processor
def inject_language():
    """템플릿에 언어 관련 함수 주입"""
    return dict(
        get_text=get_text,
        current_lang=session.get('language', 'ko')
    )

@app.route('/set-language/<language>')
def set_language(language):
    """언어 설정"""
    if language in ['ko', 'en']:
        session['language'] = language
        flash(get_text('language_changed', language), 'success')
    return redirect(request.referrer or url_for('dashboard'))

# 인증 라우트들
@app.route('/login', methods=['GET', 'POST'])
@app.route('/fresh-login', methods=['GET', 'POST'])
def login():
    """로그인"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = bool(request.form.get('remember'))
        
        if not email or not password:
            flash('이메일과 패스워드를 입력해주세요.', 'error')
            return render_template('auth/login.html')
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            if not user.active:
                flash('비활성화된 계정입니다. 관리자에게 문의하세요.', 'error')
                return render_template('auth/login.html')
            
            login_user(user, remember=remember)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('dashboard'))
        else:
            flash('이메일 또는 패스워드가 올바르지 않습니다.', 'error')
    
    return render_template('auth/login.html')

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    """로그아웃"""
    try:
        logout_user()
        session.clear()
        flash('로그아웃되었습니다.', 'success')
        return redirect(url_for('login'))
    except Exception as e:
        session.clear()
        return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """프로필 설정"""
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        department_id = request.form.get('department_id') or None
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # 입력 검증
        if not name:
            flash('이름은 필수 입력 항목입니다.', 'error')
            return render_template('profile.html', departments=Department.query.all())
        
        # 이메일 중복 확인 (자신 제외)
        existing_user = User.query.filter(User.email == email, User.id != current_user.id).first()
        if existing_user:
            flash('이미 사용중인 이메일입니다.', 'error')
            return render_template('profile.html', departments=Department.query.all())
        
        # 비밀번호 변경 요청이 있는 경우
        if new_password:
            if not current_password:
                flash('현재 비밀번호를 입력해주세요.', 'error')
                return render_template('profile.html', departments=Department.query.all())
            
            if not current_user.check_password(current_password):
                flash('현재 비밀번호가 올바르지 않습니다.', 'error')
                return render_template('profile.html', departments=Department.query.all())
            
            if new_password != confirm_password:
                flash('새 비밀번호가 일치하지 않습니다.', 'error')
                return render_template('profile.html', departments=Department.query.all())
            
            if new_password is not None and len(new_password) < 6:
                flash('비밀번호는 최소 6자 이상이어야 합니다.', 'error')
                return render_template('profile.html', departments=Department.query.all())
            
            current_user.set_password(new_password)
        
        # 프로필 정보 업데이트
        current_user.name = name
        current_user.email = email
        current_user.department_id = int(department_id) if department_id else None
        current_user.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash('프로필이 성공적으로 업데이트되었습니다.', 'success')
        return redirect(url_for('profile'))
    
    departments = Department.query.all()
    return render_template('profile.html', departments=departments)

@app.route('/register', methods=['GET', 'POST'])
def register():
    """회원가입"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        name = request.form.get('name')
        department_id = request.form.get('department_id') or None
        
        # 입력 검증
        if not all([email, password, confirm_password, name]):
            flash('모든 필드를 입력해주세요.', 'error')
            return render_template('auth/register.html', departments=Department.query.all())
        
        if password != confirm_password:
            flash('패스워드가 일치하지 않습니다.', 'error')
            return render_template('auth/register.html', departments=Department.query.all())
        
        if password and len(password) < 6:
            flash('패스워드는 최소 6자 이상이어야 합니다.', 'error')
            return render_template('auth/register.html', departments=Department.query.all())
        
        # 이메일 중복 확인
        if User.query.filter_by(email=email).first():
            flash('이미 등록된 이메일입니다.', 'error')
            return render_template('auth/register.html', departments=Department.query.all())
        
        # 새 사용자 생성
        user = User()
        user.email = email
        user.name = name
        user.department_id = department_id
        user.role = 'user'  # 기본값은 일반 사용자
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('회원가입이 완료되었습니다. 로그인해주세요.', 'success')
        return redirect(url_for('login'))
    
    departments = Department.query.all()
    return render_template('auth/register.html', departments=departments)

@app.route('/')
@login_required
def dashboard():
    """대시보드 - KPI 및 주요 지표 표시"""
    from datetime import timedelta
    # KPI 계산
    today = datetime.now().date()
    current_month = today.replace(day=1)
    last_month = (current_month - timedelta(days=1)).replace(day=1)
    
    # 이번달 수입/지출 (이체 제외)
    current_income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.amount > 0,
        Transaction.transaction_type != 'transfer',
        func.date(Transaction.transaction_date) >= current_month
    ).scalar() or 0
    
    current_expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.amount < 0,
        Transaction.transaction_type != 'transfer',
        func.date(Transaction.transaction_date) >= current_month
    ).scalar() or 0
    
    # 저번달 수입/지출 (이체 제외)
    last_income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.amount > 0,
        Transaction.transaction_type != 'transfer',
        func.date(Transaction.transaction_date) >= last_month,
        func.date(Transaction.transaction_date) < current_month
    ).scalar() or 0
    
    last_expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.amount < 0,
        Transaction.transaction_type != 'transfer',
        func.date(Transaction.transaction_date) >= last_month,
        func.date(Transaction.transaction_date) < current_month
    ).scalar() or 0
    
    # 미분류 거래 수
    unclassified_count = Transaction.query.filter_by(classification_status='pending').count()
    
    # 부서 수
    departments_count = Department.query.count()
    
    # 최근 거래 내역 (5건)
    recent_transactions = Transaction.query.order_by(desc(Transaction.transaction_date)).limit(5).all()
    
    # 최근 알림 (5건)
    recent_alerts = Alert.query.filter_by(is_read=False).order_by(desc(Alert.created_at)).limit(5).all()
    
    # 부서별 지출 현황 (최근 3개월로 확장하여 데이터 확보)
    three_months_ago = current_month - timedelta(days=90)
    
    dept_expenses_raw = db.session.query(
        Department.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.amount < 0,
        Transaction.transaction_type != 'transfer',
        func.date(Transaction.transaction_date) >= three_months_ago
    ).group_by(Department.name).all()
    
    # Convert to JSON serializable format
    dept_expenses = [{'name': row.name, 'total': float(abs(row.total or 0))} for row in dept_expenses_raw]
    
    print(f"Dashboard - Department expenses data: {dept_expenses}")
    
    return render_template('dashboard.html',
                         current_income=current_income,
                         current_expense=abs(current_expense),
                         last_income=last_income,
                         last_expense=abs(last_expense),
                         unclassified_count=unclassified_count,
                         departments_count=departments_count,
                         recent_transactions=recent_transactions,
                         recent_alerts=recent_alerts,
                         dept_expenses=dept_expenses)

@app.route('/connections')
@login_required
def connections():
    """연결 관리 - 금융기관 연결 현황 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    institutions = Institution.query.all()
    consents = Consent.query.join(Institution).all()
    
    return render_template('connections.html', 
                         institutions=institutions,
                         consents=consents)

@app.route('/connect/<institution_code>')
@login_required
def connect_institution(institution_code):
    """금융기관 연결 (OAuth 시뮬레이션)"""
    institution = Institution.query.filter_by(code=institution_code).first()
    if not institution:
        flash('지원하지 않는 금융기관입니다.', 'error')
        return redirect(url_for('connections'))
    
    # OAuth 동의 시뮬레이션
    consent = Consent()
    consent.institution_id = institution.id
    consent.consent_id = f"consent_{institution_code}_{datetime.now().timestamp()}"
    consent.status = 'active'
    consent.scope = 'account:read transaction:read'
    consent.expires_at = datetime.now() + timedelta(days=180)
    db.session.add(consent)
    db.session.commit()
    
    flash(f'{institution.name} 연결이 완료되었습니다.', 'success')
    return redirect(url_for('connections'))


@app.route('/disconnect-institution/<int:consent_id>', methods=['POST'])
@login_required
def disconnect_institution(consent_id):
    """금융기관 연결 해제"""
    try:
        consent = Consent.query.get(consent_id)
        if not consent:
            return jsonify({'success': False, 'error': '연결을 찾을 수 없습니다.'})
        
        # 연결 상태를 철회로 변경
        consent.status = 'revoked'
        consent.revoked_at = datetime.now()
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'{consent.institution.name} 연결이 해제되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/refresh-connection/<int:consent_id>', methods=['POST'])
@login_required
def refresh_connection(consent_id):
    """연결 동기화/재연결"""
    try:
        consent = Consent.query.get(consent_id)
        if not consent:
            return jsonify({'success': False, 'error': '연결을 찾을 수 없습니다.'})
        
        # 동기화 시뮬레이션
        if consent.status == 'active':
            # 만료일 연장
            consent.expires_at = datetime.now() + timedelta(days=180)
            consent.updated_at = datetime.now()
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': f'{consent.institution.name} 연결이 갱신되었습니다.'
            })
        else:
            # 비활성 상태인 경우 재연결
            consent.status = 'active'
            consent.expires_at = datetime.now() + timedelta(days=180)
            consent.updated_at = datetime.now()
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': f'{consent.institution.name} 연결이 재활성화되었습니다.'
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})



@app.route('/transactions')
@login_required
def transactions():
    """거래 내역 관리"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # per_page 값 검증 (10, 20, 25, 50, 100만 허용)
    if per_page not in [10, 20, 25, 50, 100]:
        per_page = 20
    
    # 필터 파라미터
    status = request.args.get('status', '')
    department_id = request.args.get('department_id', '')
    category_id = request.args.get('category_id', '')
    account_id = request.args.get('account_id', '')
    search = request.args.get('search', '')
    
    # 기본 쿼리 (활성 거래만)
    query = Transaction.query.filter(Transaction.is_active == True)
    
    # 필터 적용
    if status:
        query = query.filter(Transaction.classification_status == status)
    if department_id:
        query = query.filter(Transaction.department_id == department_id)
    if category_id:
        query = query.filter(Transaction.category_id == category_id)
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
    if search:
        # 대소문자 구분 없는 검색을 위해 ilike 사용
        search_pattern = f'%{search.strip()}%'
        
        # 업체, 거래처, 거래내용에서 검색
        query = query.filter(
            db.or_(
                Transaction.description.ilike(search_pattern),  # 거래내용
                Transaction.counterparty.ilike(search_pattern),  # 거래처
                Transaction.vendor.has(Vendor.name.ilike(search_pattern))  # 업체명
            )
        )
    
    # 페이지네이션
    transactions_page = query.order_by(desc(Transaction.transaction_date)).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # 필터 옵션을 위한 데이터
    departments = Department.query.all()
    categories = Category.query.all()
    accounts = Account.query.join(Institution).all()
    vendors = Vendor.query.all()  # 거래 추가 모달용
    
    # 현재 선택된 계좌 정보
    selected_account = None
    if account_id:
        selected_account = Account.query.get(account_id)
    
    return render_template('transactions.html',
                         transactions=transactions_page.items,
                         pagination=transactions_page,
                         departments=departments,
                         categories=categories,
                         accounts=accounts,
                         vendors=vendors,
                         selected_account=selected_account,
                         current_filters={
                             'status': status,
                             'department_id': department_id,
                             'category_id': category_id,
                             'account_id': account_id,
                             'search': search,
                             'per_page': per_page
                         })

@app.route('/transaction/<int:transaction_id>/details', methods=['GET'])
@login_required
def transaction_details(transaction_id):
    """거래 상세 정보 조회 (JSON)"""
    try:
        transaction = Transaction.query.get_or_404(transaction_id)
        
        transaction_data = {
            'id': transaction.id,
            'counterparty': transaction.counterparty,
            'description': transaction.description,
            'amount': float(transaction.amount),
            'transaction_date': transaction.transaction_date.isoformat(),
            'classification_status': transaction.classification_status,
            'account': {
                'institution': {'name': transaction.account.institution.name},
                'account_name': transaction.account.account_name
            }
        }
        
        return jsonify({'success': True, 'transaction': transaction_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/transaction/<int:transaction_id>/delete', methods=['POST'])
@login_required
def delete_transaction(transaction_id):
    """거래 삭제 (관리자 전용)"""
    # 관리자 권한 확인
    if not current_user.is_admin():
        return jsonify({
            'success': False,
            'error': '관리자만 거래를 삭제할 수 있습니다.'
        })
    
    try:
        transaction = Transaction.query.get(transaction_id)
        
        if not transaction:
            return jsonify({
                'success': False,
                'error': '해당 거래를 찾을 수 없습니다.'
            })
        
        # 거래를 소프트 삭제 (is_active = False)
        transaction.is_active = False
        db.session.commit()
        
        # 감사 로그 추가
        audit_log = AuditLog()
        audit_log.user_id = current_user.id
        audit_log.action = 'delete_transaction'
        audit_log.table_name = 'transactions'
        audit_log.record_id = transaction_id
        audit_log.new_values = f'거래 삭제: {transaction.description}'
        audit_log.created_at = datetime.now()
        
        db.session.add(audit_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '거래가 성공적으로 삭제되었습니다.'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'삭제 중 오류가 발생했습니다: {str(e)}'
        })

@app.route('/transaction/<int:transaction_id>/split', methods=['POST'])
@login_required
def split_transaction(transaction_id):
    """거래 분할 처리"""
    try:
        original_transaction = Transaction.query.get_or_404(transaction_id)
        
        # 분할 데이터 받기
        amounts = request.form.getlist('amounts[]')
        descriptions = request.form.getlist('descriptions[]')
        categories = request.form.getlist('categories[]')
        departments = request.form.getlist('departments[]')
        
        # 금액 합계 검증
        total_amount = sum(float(amount) for amount in amounts)
        if abs(total_amount - float(original_transaction.amount)) > 0.01:
            flash('분할된 금액의 합계가 원본 거래 금액과 일치하지 않습니다.', 'error')
            return redirect(url_for('transactions'))
        
        # 원본 거래를 비활성화
        original_transaction.is_active = False
        original_transaction.split_parent_id = None  # 이것이 분할의 부모임을 표시
        
        # 분할된 거래들 생성
        for i, amount in enumerate(amounts):
            split_transaction = Transaction()
            split_transaction.account_id = original_transaction.account_id
            split_transaction.counterparty = original_transaction.counterparty
            split_transaction.description = descriptions[i] if descriptions[i] else f"{original_transaction.description} (분할 {i+1})"
            split_transaction.amount = float(amount)
            split_transaction.transaction_date = original_transaction.transaction_date
            split_transaction.processed_date = original_transaction.processed_date
            split_transaction.transaction_type = original_transaction.transaction_type
            split_transaction.classification_status = 'manual'  # 분할된 거래는 수동분류로 설정
            split_transaction.split_parent_id = original_transaction.id  # 원본 거래 ID 저장
            
            # 카테고리와 부서 설정
            if categories[i]:
                split_transaction.category_id = int(categories[i])
            if departments[i]:
                split_transaction.department_id = int(departments[i])
            
            db.session.add(split_transaction)
        
        db.session.commit()
        flash(f'거래가 {len(amounts)}개로 성공적으로 분할되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'거래 분할 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('transactions'))

@app.route('/transaction/<int:transaction_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_transaction(transaction_id):
    """거래 내역 편집"""
    transaction = Transaction.query.get_or_404(transaction_id)
    
    if request.method == 'POST':
        try:
            # 기본 정보 업데이트
            if request.form.get('description'):
                transaction.description = request.form.get('description')
            if request.form.get('counterparty'):
                transaction.counterparty = request.form.get('counterparty')
            if request.form.get('amount'):
                amount_str = request.form.get('amount')
                if amount_str:
                    transaction.amount = float(amount_str)
            if request.form.get('transaction_date'):
                from datetime import datetime
                date_str = request.form.get('transaction_date')
                if date_str:
                    transaction.transaction_date = datetime.fromisoformat(date_str)
            
            # 분류 정보 업데이트
            transaction.category_id = request.form.get('category_id') or None
            transaction.department_id = request.form.get('department_id') or None
            transaction.vendor_id = request.form.get('vendor_id') or None
            transaction.classification_status = request.form.get('classification_status', 'manual')
            transaction.memo = request.form.get('memo') or None
            
            # 수정일시 업데이트
            from datetime import datetime
            transaction.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('거래 내역이 수정되었습니다.', 'success')
            return redirect(url_for('transactions'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    categories = Category.query.all()
    departments = Department.query.all()
    vendors = Vendor.query.all()
    
    return render_template('edit_transaction.html',
                         transaction=transaction,
                         categories=categories,
                         departments=departments,
                         vendors=vendors)

@app.route('/rules')
@login_required
def rules():
    """분류 규칙 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    rules = MappingRule.query.order_by(MappingRule.priority.desc()).all()
    categories = Category.query.all()
    departments = Department.query.all()
    vendors = Vendor.query.all()
    
    # 규칙을 딕셔너리로 변환 (JSON 직렬화를 위해)
    rules_dict = []
    for rule in rules:
        rule_dict = {
            'id': rule.id,
            'name': rule.name,
            'priority': rule.priority,
            'is_active': rule.is_active,
            'condition_type': rule.condition_type,
            'condition_field': rule.condition_field,
            'condition_value': rule.condition_value,
            'target_category_id': rule.target_category_id,
            'target_department_id': rule.target_department_id,
            'target_vendor_id': rule.target_vendor_id,
            'created_at': rule.created_at.isoformat() if rule.created_at else None
        }
        rules_dict.append(rule_dict)
    
    return render_template('rules.html',
                         rules=rules,
                         rules_dict=rules_dict,
                         categories=categories,
                         departments=departments,
                         vendors=vendors)

@app.route('/rule/add', methods=['POST'])
@login_required
def add_rule():
    """분류 규칙 추가"""
    rule = MappingRule()
    rule.name = request.form['name']
    rule.priority = int(request.form.get('priority', 0))
    rule.condition_type = request.form['condition_type']
    rule.condition_field = request.form['condition_field']
    rule.condition_value = request.form['condition_value']
    rule.target_category_id = request.form.get('target_category_id') or None
    rule.target_department_id = request.form.get('target_department_id') or None
    rule.target_vendor_id = request.form.get('target_vendor_id') or None
    
    db.session.add(rule)
    db.session.commit()
    
    flash('분류 규칙이 추가되었습니다.', 'success')
    return redirect(url_for('rules'))

@app.route('/rule/<int:rule_id>/edit', methods=['POST'])
@login_required
def edit_rule(rule_id):
    """분류 규칙 수정"""
    try:
        rule = MappingRule.query.get_or_404(rule_id)
        old_rule_active = rule.is_active
        
        # 수정 전 상태가 활성화였다면 기존 분류 취소
        if old_rule_active:
            revert_rule_classifications(rule)
        
        # 규칙 정보 업데이트
        rule.name = request.form['name']
        rule.priority = int(request.form.get('priority', 0))
        rule.condition_type = request.form['condition_type']
        rule.condition_field = request.form['condition_field']
        rule.condition_value = request.form['condition_value']
        rule.target_category_id = request.form.get('target_category_id') or None
        rule.target_department_id = request.form.get('target_department_id') or None
        rule.target_vendor_id = request.form.get('target_vendor_id') or None
        rule.is_active = 'is_active' in request.form
        
        # 수정 후에는 모든 활성 규칙을 다시 적용하여 일관성 보장
        db.session.commit()  # 규칙 수정사항 먼저 저장
        
        if rule.is_active:
            total_applied = apply_all_active_rules()
            db.session.commit()
            flash(f'분류 규칙이 수정되어 {total_applied}건의 거래가 재분류되었습니다.', 'success')
        else:
            db.session.commit()
            flash('분류 규칙이 수정되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류 규칙 수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('rules'))

def check_rule_match(rule, transaction):
    """거래가 규칙 조건에 매칭되는지 확인하는 함수"""
    match = False
    
    if rule.condition_type == 'contains':
        if rule.condition_field == 'description':
            match = rule.condition_value.lower() in (transaction.description or '').lower()
        elif rule.condition_field == 'counterparty':
            match = rule.condition_value.lower() in (transaction.counterparty or '').lower()
    
    elif rule.condition_type == 'equals':
        if rule.condition_field == 'description':
            match = (transaction.description or '').lower() == rule.condition_value.lower()
        elif rule.condition_field == 'counterparty':
            match = (transaction.counterparty or '').lower() == rule.condition_value.lower()
    
    elif rule.condition_type == 'amount_range':
        try:
            range_parts = rule.condition_value.split('-')
            if len(range_parts) == 2:
                min_amount = float(range_parts[0])
                max_amount = float(range_parts[1])
                transaction_amount = abs(transaction.amount)
                match = min_amount <= transaction_amount <= max_amount
        except:
            match = False
    
    elif rule.condition_type == 'regex':
        try:
            import re
            if rule.condition_field == 'description':
                match = bool(re.search(rule.condition_value, transaction.description or '', re.IGNORECASE))
            elif rule.condition_field == 'counterparty':
                match = bool(re.search(rule.condition_value, transaction.counterparty or '', re.IGNORECASE))
        except:
            match = False
    
    return match

def revert_rule_classifications(rule):
    """규칙에 매칭되는 모든 거래를 미분류로 되돌리는 함수 (일관성 보장)"""
    # 모든 분류된 거래 조회
    all_transactions = Transaction.query.filter(Transaction.classification_status.in_(['classified', 'manual'])).all()
    reverted_count = 0
    
    for transaction in all_transactions:
        # 해당 규칙 조건에 매칭되는 모든 거래를 미분류로 되돌림 (현재 분류와 무관)
        if check_rule_match(rule, transaction):
            transaction.classification_status = 'pending'
            transaction.category_id = None
            transaction.department_id = None
            transaction.vendor_id = None
            reverted_count += 1
    
    return reverted_count

def apply_all_active_rules():
    """모든 활성 규칙을 우선순위 순서로 적용하는 함수"""
    active_rules = MappingRule.query.filter_by(is_active=True).order_by(MappingRule.priority.desc()).all()
    total_applied = 0
    
    for rule in active_rules:
        matched_transactions = apply_rule_to_transactions(rule)
        total_applied += len(matched_transactions)
    
    return total_applied

@app.route('/rule/<int:rule_id>/toggle', methods=['POST'])
@login_required
def toggle_rule(rule_id):
    """분류 규칙 활성/비활성 토글 (일관성 보장)"""
    try:
        rule = MappingRule.query.get_or_404(rule_id)
        was_active = rule.is_active
        rule.is_active = not rule.is_active
        
        # 활성화 시: 모든 활성 규칙을 우선순위 순서로 다시 적용
        if rule.is_active and not was_active:
            db.session.commit()  # 규칙 상태 먼저 저장
            total_applied = apply_all_active_rules()
            db.session.commit()
            flash(f'규칙 "{rule.name}"이 활성화되어 {total_applied}건의 거래가 재분류되었습니다.', 'success')
        
        # 비활성화 시: 해당 규칙 매칭 거래를 미분류로 되돌린 후, 나머지 활성 규칙들 다시 적용
        elif not rule.is_active and was_active:
            db.session.commit()  # 규칙 상태 먼저 저장
            reverted_count = revert_rule_classifications(rule)
            reapplied_count = apply_all_active_rules()
            db.session.commit()
            flash(f'규칙 "{rule.name}"이 비활성화되어 {reverted_count}건의 거래가 미분류 후 {reapplied_count}건이 재분류되었습니다.', 'success')
        
        else:
            db.session.commit()
            status = "활성화" if rule.is_active else "비활성화"
            flash(f'규칙 "{rule.name}"이 {status}되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'규칙 상태 변경 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('rules'))

def apply_rule_to_transactions(rule, target_transactions=None):
    """규칙을 거래에 적용하는 공통 함수"""
    if target_transactions is None:
        # 모든 거래 대상
        target_transactions = Transaction.query.all()
    
    matched_transactions = []
    
    for transaction in target_transactions:
        match = False
        
        # 조건 타입에 따른 매칭 로직
        if rule.condition_type == 'contains':
            if rule.condition_field == 'description':
                match = rule.condition_value.lower() in (transaction.description or '').lower()
            elif rule.condition_field == 'counterparty':
                match = rule.condition_value.lower() in (transaction.counterparty or '').lower()
        
        elif rule.condition_type == 'equals':
            if rule.condition_field == 'description':
                match = (transaction.description or '').lower() == rule.condition_value.lower()
            elif rule.condition_field == 'counterparty':
                match = (transaction.counterparty or '').lower() == rule.condition_value.lower()
        
        elif rule.condition_type == 'amount_range':
            try:
                range_parts = rule.condition_value.split('-')
                if len(range_parts) == 2:
                    min_amount = float(range_parts[0])
                    max_amount = float(range_parts[1])
                    transaction_amount = abs(transaction.amount)
                    match = min_amount <= transaction_amount <= max_amount
            except:
                match = False
        
        elif rule.condition_type == 'regex':
            try:
                import re
                if rule.condition_field == 'description':
                    match = bool(re.search(rule.condition_value, transaction.description or '', re.IGNORECASE))
                elif rule.condition_field == 'counterparty':
                    match = bool(re.search(rule.condition_value, transaction.counterparty or '', re.IGNORECASE))
            except:
                match = False
        
        if match:
            # 규칙 적용
            if rule.target_category_id:
                transaction.category_id = rule.target_category_id
            if rule.target_department_id:
                transaction.department_id = rule.target_department_id
            if rule.target_vendor_id:
                transaction.vendor_id = rule.target_vendor_id
            transaction.classification_status = 'classified'
            matched_transactions.append(transaction)
    
    return matched_transactions

@app.route('/rule/<int:rule_id>/apply')
@login_required
def apply_rule(rule_id):
    """규칙 적용 실행"""
    rule = MappingRule.query.get_or_404(rule_id)
    
    # 비활성화된 규칙은 적용하지 않음
    if not rule.is_active:
        flash(f'규칙 "{rule.name}"이 비활성화 상태입니다. 먼저 활성화해주세요.', 'warning')
        return redirect(url_for('rules'))
    
    # 모든 거래에 규칙 적용
    matched_transactions = apply_rule_to_transactions(rule)
    
    db.session.commit()
    
    flash(f'{len(matched_transactions)}건의 거래가 분류되었습니다.', 'success')
    return redirect(url_for('rules'))

@app.route('/rule/<int:rule_id>/test')
@login_required
def test_rule(rule_id):
    """규칙 테스트 - 실제 거래 데이터에서 매칭되는 거래들 찾기"""
    try:
        rule = MappingRule.query.get_or_404(rule_id)
        
        # 모든 거래 조회 (최근 100건으로 제한)
        transactions = Transaction.query.order_by(Transaction.transaction_date.desc()).limit(100).all()
        
        matched_transactions = []
        
        for transaction in transactions:
            match = False
            
            # 조건 타입에 따른 매칭 로직
            if rule.condition_type == 'contains':
                if rule.condition_field == 'description':
                    match = rule.condition_value.lower() in (transaction.description or '').lower()
                elif rule.condition_field == 'counterparty':
                    match = rule.condition_value.lower() in (transaction.counterparty or '').lower()
            
            elif rule.condition_type == 'equals':
                if rule.condition_field == 'description':
                    match = (transaction.description or '').lower() == rule.condition_value.lower()
                elif rule.condition_field == 'counterparty':
                    match = (transaction.counterparty or '').lower() == rule.condition_value.lower()
            
            elif rule.condition_type == 'amount_range':
                try:
                    # 금액 범위 조건 (예: "1000-5000")
                    range_parts = rule.condition_value.split('-')
                    if len(range_parts) == 2:
                        min_amount = float(range_parts[0])
                        max_amount = float(range_parts[1])
                        transaction_amount = abs(transaction.amount)
                        match = min_amount <= transaction_amount <= max_amount
                except:
                    match = False
            
            elif rule.condition_type == 'regex':
                try:
                    import re
                    if rule.condition_field == 'description':
                        match = bool(re.search(rule.condition_value, transaction.description or '', re.IGNORECASE))
                    elif rule.condition_field == 'counterparty':
                        match = bool(re.search(rule.condition_value, transaction.counterparty or '', re.IGNORECASE))
                except:
                    match = False
            
            if match:
                matched_transactions.append({
                    'date': transaction.transaction_date.strftime('%Y-%m-%d') if transaction.transaction_date else '',
                    'description': transaction.description or '',
                    'counterparty': transaction.counterparty or '',
                    'amount': transaction.amount,
                    'current_category': transaction.category.name if transaction.category else '미분류',
                    'target_category': rule.target_category.name if rule.target_category else '',
                    'target_department': rule.target_department.name if rule.target_department else '',
                    'target_vendor': rule.target_vendor.name if rule.target_vendor else ''
                })
        
        result = {
            'rule_name': rule.name,
            'matched_count': len(matched_transactions),
            'transactions': matched_transactions[:10],  # 최대 10건만 표시
            'condition_info': {
                'type': rule.condition_type,
                'field': rule.condition_field,
                'value': rule.condition_value
            }
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/reports')
@login_required
def reports():
    """보고서"""
    # 현금흐름 차트 데이터 (최근 12개월)
    end_date = date.today()
    start_date = end_date.replace(year=end_date.year - 1)
    
    # 월별 현금흐름 (수입과 지출 분리)
    monthly_income_raw = db.session.query(
        extract('year', Transaction.transaction_date).label('year'),
        extract('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('total')
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.amount > 0
    ).group_by(
        extract('year', Transaction.transaction_date),
        extract('month', Transaction.transaction_date)
    ).order_by('year', 'month').all()
    
    monthly_expense_raw = db.session.query(
        extract('year', Transaction.transaction_date).label('year'),
        extract('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('total')
    ).filter(
        Transaction.transaction_date >= start_date,
        Transaction.amount < 0
    ).group_by(
        extract('year', Transaction.transaction_date),
        extract('month', Transaction.transaction_date)
    ).order_by('year', 'month').all()
    
    # 수입과 지출 데이터를 합치기
    income_dict = {(int(row.year), int(row.month)): float(row.total or 0) for row in monthly_income_raw}
    expense_dict = {(int(row.year), int(row.month)): float(row.total or 0) for row in monthly_expense_raw}
    
    # 모든 월에 대해 수입과 지출 데이터 생성
    all_months = set(income_dict.keys()) | set(expense_dict.keys())
    monthly_flow = []
    for year, month in sorted(all_months):
        income = income_dict.get((year, month), 0)
        expense = expense_dict.get((year, month), 0)
        monthly_flow.append({
            'year': year,
            'month': month,
            'income': income,
            'expense': expense,
            'total': income + expense
        })
    
    # 부서별 지출 현황 (이번달)
    current_month = end_date.replace(day=1)
    dept_spending_raw = db.session.query(
        Department.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.amount < 0,
        Transaction.transaction_date >= current_month
    ).group_by(Department.name).all()
    
    # Convert to JSON serializable format
    dept_spending = [{'name': row.name, 'total': float(abs(row.total or 0))} for row in dept_spending_raw]
    
    # 상위 거래처 (이번달)
    top_vendors = db.session.query(
        Vendor.name,
        func.sum(Transaction.amount).label('total'),
        func.count(Transaction.id).label('count')
    ).join(Transaction).filter(
        Transaction.amount < 0,
        Transaction.transaction_date >= current_month
    ).group_by(Vendor.name).order_by(desc('total')).limit(10).all()
    
    return render_template('reports.html',
                         monthly_flow=monthly_flow,
                         dept_spending=dept_spending,
                         top_vendors=top_vendors)


@app.route('/reports/data')
@login_required
def reports_data():
    """새로운 6가지 리포트 데이터 AJAX 요청 처리"""
    try:
        from datetime import datetime, date
        from models import Transaction, Account, Category, Department, Vendor
        from sqlalchemy import func, extract, case
        
        # 요청 파라미터 받기
        report_type = request.args.get('report_type', 'pl')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        print(f"Reports data request - Type: {report_type}, Start: {start_date_str}, End: {end_date_str}")
        
        # 날짜 파싱
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError as e:
                return jsonify({'success': False, 'error': f'날짜 형식 오류: {str(e)}'})
        else:
            # 기본값: 올해 전체
            today = date.today()
            start_date = date(today.year, 1, 1)
            end_date = today
        
        data = {}
        
        if report_type == 'pl':
            # 손익계산서 - 수익과 비용 분석
            revenue_data = db.session.query(
                Category.name,
                func.sum(Transaction.amount).label('total')
            ).join(Transaction).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date,
                Transaction.amount > 0
            ).group_by(Category.name).all()
            
            cost_data = db.session.query(
                Category.name,
                func.sum(Transaction.amount).label('total')
            ).join(Transaction).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date,
                Transaction.amount < 0
            ).group_by(Category.name).all()
            
            pl_data = []
            for row in revenue_data:
                pl_data.append({
                    'name': row.name,
                    'type': 'revenue',
                    'amount': float(row.total or 0)
                })
            
            for row in cost_data:
                pl_data.append({
                    'name': row.name,
                    'type': 'cost',
                    'amount': float(abs(row.total or 0))
                })
            
            data['pl'] = pl_data
            
        elif report_type == 'cashflow':
            # 현금흐름표 - 월별 현금흐름
            monthly_flow_raw = db.session.query(
                extract('year', Transaction.transaction_date).label('year'),
                extract('month', Transaction.transaction_date).label('month'),
                func.sum(case((Transaction.amount > 0, Transaction.amount), else_=0)).label('income'),
                func.sum(case((Transaction.amount < 0, Transaction.amount), else_=0)).label('expense')
            ).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).group_by(
                extract('year', Transaction.transaction_date),
                extract('month', Transaction.transaction_date)
            ).order_by('year', 'month').all()
            
            cashflow_data = []
            for row in monthly_flow_raw:
                income = float(row.income or 0)
                expense = float(row.expense or 0)
                cashflow_data.append({
                    'period': f"{int(row.year)}-{str(int(row.month)).zfill(2)}",
                    'income': income,
                    'expense': abs(expense),
                    'net': income + expense
                })
            
            data['cashflow'] = cashflow_data
            
        elif report_type == 'budget':
            # 예산 vs 실적
            # 부서별 예산 데이터 (없으면 0으로 처리)
            dept_budget_raw = db.session.query(Department.name).all()
            
            budget_data = []
            for dept in dept_budget_raw:
                # 실제 지출 계산
                actual_spent = db.session.query(
                    func.sum(Transaction.amount).label('total')
                ).join(Department).filter(
                    Department.name == dept.name,
                    Transaction.transaction_date >= start_date,
                    Transaction.transaction_date <= end_date,
                    Transaction.amount < 0
                ).scalar() or 0
                
                # 임시 예산 설정 (실제로는 Budget 테이블에서 가져와야 함)
                budget_amount = 1000000  # 100만원 기본 예산
                
                budget_data.append({
                    'department': dept.name,
                    'budget': budget_amount,
                    'actual': abs(float(actual_spent)),
                    'variance': budget_amount - abs(float(actual_spent))
                })
            
            data['budget'] = budget_data
            
        elif report_type == 'department':
            # 부서별 손익 분석
            dept_analysis_raw = db.session.query(
                Department.name,
                func.sum(case((Transaction.amount > 0, Transaction.amount), else_=0)).label('revenue'),
                func.sum(case((Transaction.amount < 0, Transaction.amount), else_=0)).label('cost'),
                func.count(Transaction.id).label('count')
            ).join(Transaction).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).group_by(Department.name).all()
            
            dept_data = []
            for row in dept_analysis_raw:
                revenue = float(row.revenue or 0)
                cost = abs(float(row.cost or 0))
                dept_data.append({
                    'name': row.name,
                    'revenue': revenue,
                    'cost': cost,
                    'profit': revenue - cost,
                    'count': row.count
                })
            
            data['department'] = dept_data
            
        elif report_type == 'vendor':
            # 거래처별 분석
            vendor_analysis_raw = db.session.query(
                Vendor.name,
                func.sum(Transaction.amount).label('total'),
                func.count(Transaction.id).label('count')
            ).join(Transaction).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).group_by(Vendor.name).order_by(func.abs(func.sum(Transaction.amount)).desc()).limit(20).all()
            
            vendor_data = [{
                'name': row.name,
                'total': float(row.total or 0),
                'count': row.count
            } for row in vendor_analysis_raw]
            
            data['vendor'] = vendor_data
            
        elif report_type == 'category':
            # 카테고리별 분석
            category_analysis_raw = db.session.query(
                Category.name,
                func.sum(Transaction.amount).label('total'),
                func.count(Transaction.id).label('count')
            ).join(Transaction).filter(
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).group_by(Category.name).order_by(func.abs(func.sum(Transaction.amount)).desc()).all()
            
            category_data = [{
                'name': row.name,
                'total': float(row.total or 0),
                'count': row.count
            } for row in category_analysis_raw]
            
            data['category'] = category_data
        
        print(f"Reports data response: {data}")
        return jsonify({'success': True, 'data': data})
        
    except Exception as e:
        print(f"Reports data error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/settings')
@login_required
def settings():
    """설정 및 알림 - alerts.html로 리다이렉트"""
    return redirect(url_for('alerts'))

@app.route('/alerts')
@login_required
def alerts():
    """알림 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import AlertSetting
    alerts = Alert.query.order_by(desc(Alert.created_at)).limit(50).all()
    alert_settings = AlertSetting.query.order_by(desc(AlertSetting.created_at)).all()
    
    # 기존 시스템 설정값들 조회
    current_settings = {
        'budgetAlert': False,
        'budgetThreshold': '80',
        'budgetFrequency': 'weekly',
        'contractAlert': False,
        'contractDays': '30',
        'contractFrequency': 'daily',
        'anomalyAlert': False,
        'anomalyAmount': '1000000',
        'immediateAlert': True
    }
    
    # 데이터베이스에서 현재 설정 불러오기
    for setting in alert_settings:
        if setting.alert_type == 'budget' and setting.is_active:
            current_settings['budgetAlert'] = True
            current_settings['budgetThreshold'] = setting.condition_value
        elif setting.alert_type == 'contract' and setting.is_active:
            current_settings['contractAlert'] = True
            current_settings['contractDays'] = setting.condition_value
        elif setting.alert_type == 'anomaly' and setting.is_active:
            current_settings['anomalyAlert'] = True
            # condition_value는 "10000,9999999999" 형태이므로 첫 번째 값만 가져옴
            if ',' in setting.condition_value:
                current_settings['anomalyAmount'] = setting.condition_value.split(',')[0]
    
    # 세션에서도 설정값 확인 (최신 설정 우선)
    session_settings = session.get('alert_settings', {})
    if session_settings:
        current_settings.update(session_settings)
    
    # current_settings가 비어있지 않은지 확인
    if not current_settings:
        current_settings = {
            'budgetAlert': False,
            'budgetThreshold': '80',
            'budgetFrequency': 'weekly',
            'contractAlert': False,
            'contractDays': '30',
            'contractFrequency': 'daily',
            'anomalyAlert': False,
            'anomalyAmount': '1000000',
            'immediateAlert': True
        }
    
    return render_template('alerts.html', alerts=alerts, alert_settings=alert_settings, current_settings=current_settings)


@app.route('/alerts/export')
@login_required
def export_alerts():
    """알림 내역 내보내기"""
    try:
        from datetime import date, datetime
        import csv
        import io
        from flask import make_response
        
        # 파라미터 가져오기
        export_format = request.args.get('format', 'csv').lower()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        alert_type = request.args.get('alert_type', 'all')
        
        # 날짜 처리
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # 기본값: 지난 6개월
            end_date = datetime.now()
            start_date = end_date.replace(month=end_date.month - 6 if end_date.month > 6 else end_date.month + 6, 
                                        year=end_date.year if end_date.month > 6 else end_date.year - 1)
        
        # Alert 데이터 조회
        query = Alert.query.filter(
            Alert.created_at >= start_date,
            Alert.created_at <= end_date
        )
        
        if alert_type != 'all':
            query = query.filter(Alert.alert_type == alert_type)
        
        alerts = query.order_by(Alert.created_at.desc()).all()
        
        if export_format == 'csv':
            return export_alerts_csv(alerts, start_date, end_date)
        elif export_format == 'excel':
            return export_alerts_excel(alerts, start_date, end_date)
        else:
            flash('지원하지 않는 내보내기 형식입니다.', 'error')
            return redirect(url_for('alerts'))
            
    except Exception as e:
        flash(f'알림 내역 내보내기 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('alerts'))


def export_alerts_csv(alerts, start_date, end_date):
    """CSV 형식으로 알림 데이터 내보내기"""
    import csv
    import io
    from flask import make_response
    
    # UTF-8 BOM을 포함한 BytesIO 사용
    output = io.BytesIO()
    
    # UTF-8 BOM 추가 (Excel에서 한글 정상 표시를 위해)
    output.write('\ufeff'.encode('utf-8'))
    
    # CSV 데이터를 문자열로 생성
    csv_data = io.StringIO()
    writer = csv.writer(csv_data)
    
    # 헤더 작성
    writer.writerow([
        '생성일시', '제목', '메시지', '알림 유형', '심각도', '읽음 상태', '관련 테이블', '관련 ID'
    ])
    
    # 데이터 작성
    for alert in alerts:
        writer.writerow([
            alert.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            alert.title,
            alert.message,
            get_alert_type_name(alert.alert_type),
            get_severity_name(alert.severity),
            '읽음' if alert.is_read else '미읽음',
            alert.related_table or '',
            alert.related_id or ''
        ])
    
    # UTF-8로 인코딩해서 BytesIO에 추가
    output.write(csv_data.getvalue().encode('utf-8'))
    output.seek(0)
    
    # 응답 생성
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    filename = f'alerts_history_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.csv'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


def export_alerts_excel(alerts, start_date, end_date):
    """Excel 형식으로 알림 데이터 내보내기"""
    import io
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "알림 내역"
    
    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center")
    
    # 헤더 작성
    headers = ['생성일시', '제목', '메시지', '알림 유형', '심각도', '읽음 상태', '관련 테이블', '관련 ID']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row, alert in enumerate(alerts, 2):
        ws.cell(row=row, column=1, value=alert.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=alert.title)
        ws.cell(row=row, column=3, value=alert.message)
        ws.cell(row=row, column=4, value=get_alert_type_name(alert.alert_type))
        ws.cell(row=row, column=5, value=get_severity_name(alert.severity))
        ws.cell(row=row, column=6, value='읽음' if alert.is_read else '미읽음')
        ws.cell(row=row, column=7, value=alert.related_table or '')
        ws.cell(row=row, column=8, value=alert.related_id or '')
        
        # 미읽음 알림 강조
        if not alert.is_read:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
    
    # 컬럼 너비 조정
    column_widths = [20, 25, 40, 15, 10, 10, 15, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 파일 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    filename = f'alerts_history_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


def get_alert_type_name(alert_type):
    """알림 유형 한글명 반환"""
    type_names = {
        'budget': '예산',
        'contract': '계약',
        'anomaly': '이상거래',
        'custom': '사용자 정의',
        'system': '시스템'
    }
    return type_names.get(alert_type, alert_type)


def get_severity_name(severity):
    """심각도 한글명 반환"""
    severity_names = {
        'info': '정보',
        'warning': '경고',
        'error': '오류',
        'critical': '심각'
    }
    return severity_names.get(severity, severity)

@app.route('/audit')
@login_required
def audit():
    """감사 로그"""
    return render_template('audit.html')

@app.route('/init_data')
@login_required
def init_data():
    """샘플 데이터 초기화"""
    init_sample_data()
    flash('샘플 데이터가 생성되었습니다.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/alert/<int:alert_id>/read')
@login_required
def mark_alert_read(alert_id):
    """알림 읽음 처리"""
    alert = Alert.query.get_or_404(alert_id)
    alert.is_read = True
    db.session.commit()
    
    return redirect(url_for('alerts'))

@app.route('/alerts/<int:alert_id>/delete', methods=['POST'])
@login_required
def delete_alert(alert_id):
    """알림 삭제"""
    try:
        alert = Alert.query.get(alert_id)
        
        if not alert:
            return jsonify({
                'success': False,
                'error': '해당 알림을 찾을 수 없습니다.'
            })
        
        # 알림 삭제
        alert_title = alert.title
        db.session.delete(alert)
        db.session.commit()
        
        # 감사 로그 추가
        audit_log = AuditLog()
        audit_log.user_id = current_user.id
        audit_log.action = 'delete_alert'
        audit_log.table_name = 'alerts'
        audit_log.record_id = alert_id
        audit_log.new_values = f'알림 삭제: {alert_title}'
        audit_log.created_at = datetime.now()
        
        db.session.add(audit_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '알림이 성공적으로 삭제되었습니다.'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'삭제 중 오류가 발생했습니다: {str(e)}'
        })

@app.route('/alerts/settings', methods=['POST'])
@login_required
def save_alert_settings():
    """알림 설정 저장"""
    try:
        from models import AlertSetting
        settings_data = request.get_json()
        
        # 기존 설정 삭제 (시스템 기본 알림 설정들)
        existing_settings = AlertSetting.query.filter(
            AlertSetting.alert_type.in_(['budget', 'contract', 'anomaly'])
        ).all()
        for setting in existing_settings:
            db.session.delete(setting)
        
        # 예산 알림 설정
        if settings_data.get('budgetAlert'):
            budget_setting = AlertSetting()
            budget_setting.name = "예산 초과 알림"
            budget_setting.alert_type = "budget"
            budget_setting.condition = f"budget threshold {settings_data.get('budgetThreshold', '80')}"
            budget_setting.condition_type = "percentage"
            budget_setting.condition_field = "budget_usage"
            budget_setting.condition_value = settings_data.get('budgetThreshold', '80')
            budget_setting.severity = "warning"
            budget_setting.is_active = True
            db.session.add(budget_setting)
        
        # 계약 만료 알림 설정
        if settings_data.get('contractAlert'):
            contract_setting = AlertSetting()
            contract_setting.name = "계약 만료 알림"
            contract_setting.alert_type = "contract"
            contract_setting.condition = f"contract expires in {settings_data.get('contractDays', '30')} days"
            contract_setting.condition_type = "date_range"
            contract_setting.condition_field = "end_date"
            contract_setting.condition_value = settings_data.get('contractDays', '30')
            contract_setting.severity = "info"
            contract_setting.is_active = True
            db.session.add(contract_setting)
        
        # 이상거래 알림 설정
        if settings_data.get('anomalyAlert'):
            anomaly_setting = AlertSetting()
            anomaly_setting.name = "이상거래 알림"
            anomaly_setting.alert_type = "anomaly"
            anomaly_setting.condition = f"amount >= {settings_data.get('anomalyAmount', '1000000')}"
            anomaly_setting.condition_type = "amount_range"
            anomaly_setting.condition_field = "amount"
            # 설정된 금액 이상의 거래를 감지하기 위해 최대값을 큰 수로 설정
            anomaly_amount = settings_data.get('anomalyAmount', '1000000')
            anomaly_setting.condition_value = f"{anomaly_amount},9999999999"
            anomaly_setting.severity = "warning"
            anomaly_setting.is_active = True
            db.session.add(anomaly_setting)
        
        # 사용자 정의 알림 처리
        if 'customAlerts' in settings_data:
            custom_alerts = settings_data['customAlerts']
            
            # 기존 사용자 정의 알림 삭제
            existing_custom = AlertSetting.query.filter(
                AlertSetting.alert_type == 'custom'
            ).all()
            for setting in existing_custom:
                db.session.delete(setting)
            
            # 거래처 포함 알림
            if custom_alerts.get('counterparty', {}).get('enabled') and custom_alerts['counterparty'].get('keyword'):
                counterparty_setting = AlertSetting()
                counterparty_setting.name = f"거래처 포함: {custom_alerts['counterparty']['keyword']}"
                counterparty_setting.alert_type = "custom"
                counterparty_setting.condition = f"counterparty contains {custom_alerts['counterparty']['keyword']}"
                counterparty_setting.condition_type = "contains"
                counterparty_setting.condition_field = "counterparty"
                counterparty_setting.condition_value = custom_alerts['counterparty']['keyword']
                counterparty_setting.severity = custom_alerts['counterparty'].get('severity', 'warning')
                counterparty_setting.channel = custom_alerts['counterparty'].get('channel', 'app')
                counterparty_setting.is_active = True
                db.session.add(counterparty_setting)
            
            # 거래내용 일치 알림
            if custom_alerts.get('description', {}).get('enabled') and custom_alerts['description'].get('keyword'):
                description_setting = AlertSetting()
                description_setting.name = f"거래내용 일치: {custom_alerts['description']['keyword']}"
                description_setting.alert_type = "custom"
                description_setting.condition = f"description equals {custom_alerts['description']['keyword']}"
                description_setting.condition_type = "equals"
                description_setting.condition_field = "description"
                description_setting.condition_value = custom_alerts['description']['keyword']
                description_setting.severity = custom_alerts['description'].get('severity', 'info')
                description_setting.channel = custom_alerts['description'].get('channel', 'app')
                description_setting.is_active = True
                db.session.add(description_setting)
            
            # 금액 범위 알림
            if custom_alerts.get('amount', {}).get('enabled') and custom_alerts['amount'].get('value1'):
                amount_setting = AlertSetting()
                amount_condition = custom_alerts['amount']['condition']
                value1 = custom_alerts['amount']['value1']
                value2 = custom_alerts['amount'].get('value2', '')
                
                if amount_condition == 'greater':
                    amount_setting.name = f"금액 초과: {int(value1):,}원"
                    amount_setting.condition = f"amount > {value1}"
                    amount_setting.condition_value = f"{value1},9999999999"
                elif amount_condition == 'less':
                    amount_setting.name = f"금액 미만: {int(value1):,}원"
                    amount_setting.condition = f"amount < {value1}"
                    amount_setting.condition_value = f"0,{value1}"
                elif amount_condition == 'between' and value2:
                    amount_setting.name = f"금액 범위: {int(value1):,}~{int(value2):,}원"
                    amount_setting.condition = f"amount range {value1},{value2}"
                    amount_setting.condition_value = f"{value1},{value2}"
                
                amount_setting.alert_type = "custom"
                amount_setting.condition_type = "amount_range"
                amount_setting.condition_field = "amount"
                amount_setting.severity = custom_alerts['amount'].get('severity', 'warning')
                amount_setting.channel = custom_alerts['amount'].get('channel', 'app')
                amount_setting.is_active = True
                db.session.add(amount_setting)
            
            # 고급 패턴 알림
            if custom_alerts.get('advanced', {}).get('enabled') and custom_alerts['advanced'].get('pattern'):
                advanced_setting = AlertSetting()
                advanced_setting.name = f"고급 패턴: {custom_alerts['advanced']['pattern'][:20]}..."
                advanced_setting.alert_type = "custom"
                advanced_setting.condition = f"{custom_alerts['advanced']['field']} regex {custom_alerts['advanced']['pattern']}"
                advanced_setting.condition_type = "regex"
                advanced_setting.condition_field = custom_alerts['advanced']['field']
                advanced_setting.condition_value = custom_alerts['advanced']['pattern']
                advanced_setting.severity = custom_alerts['advanced'].get('severity', 'critical')
                advanced_setting.channel = custom_alerts['advanced'].get('channel', 'app')
                advanced_setting.is_active = True
                db.session.add(advanced_setting)
        
        db.session.commit()
        
        # 세션에도 저장 (호환성 유지)
        session['alert_settings'] = settings_data
        
        return jsonify({'success': True, 'message': '알림 설정이 저장되었습니다.'})
    except Exception as e:
        db.session.rollback()
        print(f"Alert settings save error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/alerts/create-custom', methods=['POST'])
@login_required
def create_custom_alert():
    """사용자 정의 알림 생성 (중복 검사 포함)"""
    try:
        from models import AlertSetting
        
        data = request.get_json()
        alert_type = data.get('type')
        check_value = data.get('checkValue')
        alert_data = data.get('alertData')
        
        if not all([alert_type, check_value, alert_data]):
            return jsonify({'success': False, 'error': '필수 데이터가 누락되었습니다.'}), 400
        
        # 중복 검사
        is_duplicate = False
        duplicate_message = ""
        
        existing_alert = AlertSetting.query.filter_by(
            alert_type='custom',
            condition_type=alert_data['condition_type'],
            condition_field=alert_data['condition_field'],
            condition_value=alert_data['condition_value']
        ).first()
        
        if existing_alert:
            is_duplicate = True
            duplicate_message = f"동일한 조건의 알림이 이미 존재합니다: {existing_alert.name}"
        
        # 새 알림 생성 (중복이 있어도 생성)
        new_alert = AlertSetting()
        new_alert.name = alert_data['name']
        new_alert.alert_type = 'custom'
        new_alert.condition = alert_data['condition']
        new_alert.condition_type = alert_data['condition_type']
        new_alert.condition_field = alert_data['condition_field']
        new_alert.condition_value = alert_data['condition_value']
        new_alert.severity = alert_data.get('severity', 'info')
        new_alert.channel = alert_data.get('channel', 'app')
        new_alert.is_active = True
        
        db.session.add(new_alert)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'isDuplicate': is_duplicate,
            'message': duplicate_message,
            'alertId': new_alert.id
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Custom alert creation error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/alerts/create-system', methods=['POST'])
@login_required
def create_system_alert():
    """시스템 알림 생성 (중복 검사 포함)"""
    try:
        from models import AlertSetting
        
        data = request.get_json()
        alert_type = data.get('type')
        check_value = data.get('checkValue')
        alert_data = data.get('alertData')
        
        if not all([alert_type, check_value, alert_data]):
            return jsonify({'success': False, 'error': '필수 데이터가 누락되었습니다.'}), 400
        
        # 중복 검사
        is_duplicate = False
        duplicate_message = ""
        
        existing_alert = AlertSetting.query.filter_by(
            alert_type=alert_data['alert_type'],
            condition_type=alert_data['condition_type'],
            condition_field=alert_data['condition_field'],
            condition_value=alert_data['condition_value']
        ).first()
        
        if existing_alert:
            is_duplicate = True
            duplicate_message = f"동일한 조건의 알림이 이미 존재합니다: {existing_alert.name}"
        
        # 새 시스템 알림 생성 (중복이 있어도 생성)
        new_alert = AlertSetting()
        new_alert.name = alert_data['name']
        new_alert.alert_type = alert_data['alert_type']
        new_alert.condition = alert_data['condition']
        new_alert.condition_type = alert_data['condition_type']
        new_alert.condition_field = alert_data['condition_field']
        new_alert.condition_value = alert_data['condition_value']
        new_alert.severity = alert_data.get('severity', 'warning')
        new_alert.channel = alert_data.get('channel', 'app')
        new_alert.is_active = True
        
        # 추가적인 시스템 알림 속성들
        if alert_data.get('frequency'):
            new_alert.frequency = alert_data['frequency']
        if alert_data.get('immediate'):
            new_alert.immediate = alert_data['immediate']
        
        db.session.add(new_alert)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'isDuplicate': is_duplicate,
            'message': duplicate_message,
            'alertId': new_alert.id
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"System alert creation error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/alerts/add', methods=['POST'])
@login_required  
def add_alert_setting():
    """새 알림 설정 추가"""
    try:
        from utils import parse_alert_condition
        from models import AlertSetting
        
        # 폼 데이터 받기
        name = request.form.get('name')
        alert_type = request.form.get('type')
        condition = request.form.get('condition')
        severity = request.form.get('severity')
        channel = request.form.get('channel')
        
        if not name or not condition:
            flash('알림명과 조건은 필수입니다.', 'error')
            return redirect(url_for('alerts'))
        
        # 조건 파싱
        parsed = parse_alert_condition(condition)
        
        # 새 알림 설정 생성
        alert_setting = AlertSetting()
        alert_setting.name = name
        alert_setting.alert_type = alert_type or 'custom'
        alert_setting.condition = condition
        alert_setting.severity = severity or 'info'
        alert_setting.channel = channel or 'system'
        
        if parsed:
            alert_setting.condition_type = parsed['type']
            alert_setting.condition_field = parsed['field']
            alert_setting.condition_value = parsed['value']
        
        db.session.add(alert_setting)
        db.session.commit()
        
        flash(f'알림 설정 "{name}"이 추가되었습니다.', 'success')
        return redirect(url_for('alerts'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'알림 설정 추가 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('alerts'))

@app.route('/alerts/settings/<int:setting_id>/toggle', methods=['POST'])
@login_required
def toggle_alert_setting(setting_id):
    """알림 설정 활성화/비활성화"""
    try:
        from models import AlertSetting
        setting = AlertSetting.query.get_or_404(setting_id)
        setting.is_active = not setting.is_active
        db.session.commit()
        
        status = '활성화' if setting.is_active else '비활성화'
        return jsonify({'success': True, 'message': f'알림 설정이 {status}되었습니다.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/alerts/settings/<int:setting_id>/delete', methods=['POST'])
@login_required
def delete_alert_setting(setting_id):
    """알림 설정 삭제"""
    try:
        from models import AlertSetting
        setting = AlertSetting.query.get_or_404(setting_id)
        setting_name = setting.name
        db.session.delete(setting)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'알림 설정 "{setting_name}"이 삭제되었습니다.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/alerts/settings/edit', methods=['POST'])
@login_required
def edit_alert_setting():
    """알림 설정 수정"""
    try:
        from models import AlertSetting
        from utils import parse_alert_condition, generate_condition_from_type
        
        setting_id = request.form.get('setting_id')
        name = request.form.get('name')
        condition_type = request.form.get('condition_type')
        condition_value = request.form.get('condition_value')
        condition = request.form.get('condition')
        severity = request.form.get('severity', 'info')
        channel = request.form.get('channel', 'app')
        
        print(f"Edit alert setting - ID: {setting_id}, Name: {name}, Type: {condition_type}, Value: {condition_value}, Condition: {condition}")
        
        if not setting_id or not name:
            flash('필수 정보가 누락되었습니다.', 'error')
            return redirect(url_for('alerts'))
        
        setting = AlertSetting.query.get_or_404(setting_id)
        
        # 조건 생성 로직
        final_condition = condition
        if not final_condition and condition_type and condition_value:
            final_condition = generate_condition_from_type(condition_type, condition_value)
        
        # 조건이 여전히 없으면 기본값 설정
        if not final_condition:
            final_condition = f"description contains {name}"
        
        # 조건 파싱하여 구조화된 필드 업데이트
        parsed = parse_alert_condition(final_condition)
        
        setting.name = name
        setting.condition = final_condition
        setting.severity = severity
        setting.channel = channel
        
        if parsed:
            setting.condition_type = parsed['type']
            setting.condition_field = parsed['field']
            setting.condition_value = parsed['value']
        
        db.session.commit()
        
        flash(f'알림 설정 "{name}"이 수정되었습니다.', 'success')
        return redirect(url_for('alerts'))
        
    except Exception as e:
        db.session.rollback()
        print(f"Edit alert setting error: {str(e)}")
        flash(f'알림 설정 수정 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('alerts'))

@app.route('/contracts')
@login_required
def contracts():
    """계약 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import Contract, Vendor, Department, Category
    contracts = Contract.query.order_by(desc(Contract.created_at)).all()
    vendors = Vendor.query.all()
    departments = Department.query.all()
    categories = Category.query.all()
    return render_template('contracts.html', contracts=contracts, vendors=vendors, departments=departments, categories=categories)

@app.route('/contracts/add', methods=['POST'])
@login_required
def add_contract():
    """계약 추가"""
    try:
        from models import Contract
        
        name = request.form.get('name')
        vendor_id = request.form.get('vendor_id')
        department_id = request.form.get('department_id')
        contract_amount = request.form.get('contract_amount')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        description = request.form.get('description', '')
        
        if not all([name, vendor_id, department_id, contract_amount, start_date_str, end_date_str]):
            flash('필수 정보가 누락되었습니다.', 'error')
            return redirect(url_for('contracts'))
        
        # 계약금액에서 콤마 제거 후 숫자 변환
        clean_amount = contract_amount.replace(',', '')
        if not clean_amount.replace('.', '').isdigit():
            flash('올바른 계약금액을 입력해주세요.', 'error')
            return redirect(url_for('contracts'))
        
        # 날짜 변환
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        contract = Contract(
            name=name,
            vendor_id=int(vendor_id),
            department_id=int(department_id),
            contract_amount=float(clean_amount),
            start_date=start_date,
            end_date=end_date,
            description=description
        )
        
        db.session.add(contract)
        db.session.commit()
        
        flash(f'계약 "{name}"이 추가되었습니다.', 'success')
        return redirect(url_for('contracts'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'계약 추가 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('contracts'))

@app.route('/contracts/<int:contract_id>/edit', methods=['POST'])
@login_required
def edit_contract(contract_id):
    """계약 수정"""
    try:
        from models import Contract
        
        contract = Contract.query.get_or_404(contract_id)
        
        contract.name = request.form.get('name', contract.name)
        contract.vendor_id = int(request.form.get('vendor_id', contract.vendor_id))
        contract.department_id = int(request.form.get('department_id', contract.department_id))
        
        # 계약금액에서 콤마 제거 후 숫자 변환
        contract_amount_str = request.form.get('contract_amount', str(contract.contract_amount))
        clean_amount = contract_amount_str.replace(',', '')
        contract.contract_amount = float(clean_amount)
        
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        
        if start_date_str:
            from datetime import datetime
            contract.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        
        if end_date_str:
            from datetime import datetime
            contract.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        contract.description = request.form.get('description', contract.description)
        contract.status = request.form.get('status', contract.status)
        
        db.session.commit()
        
        flash(f'계약 "{contract.name}"이 수정되었습니다.', 'success')
        return redirect(url_for('contracts'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'계약 수정 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('contracts'))

@app.route('/reports/export')
@login_required
def export_reports():
    """리포트 내보내기"""
    try:
        from datetime import date, datetime
        import io
        from flask import make_response
        
        # 파라미터 가져오기
        report_type = request.args.get('type', 'cashflow')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        period = request.args.get('period', 'monthly')
        export_format = request.args.get('format', 'pdf')
        
        # 날짜 처리
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = date.today()
            start_date = end_date.replace(year=end_date.year - 1)
        
        # 데이터 조회
        from models import Transaction, Department, Vendor
        from sqlalchemy import func, extract
        
        # 월별 현금흐름 데이터
        monthly_income_raw = db.session.query(
            extract('year', Transaction.transaction_date).label('year'),
            extract('month', Transaction.transaction_date).label('month'),
            func.sum(Transaction.amount).label('total')
        ).filter(
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
            Transaction.amount > 0
        ).group_by(
            extract('year', Transaction.transaction_date),
            extract('month', Transaction.transaction_date)
        ).order_by('year', 'month').all()
        
        monthly_expense_raw = db.session.query(
            extract('year', Transaction.transaction_date).label('year'),
            extract('month', Transaction.transaction_date).label('month'),
            func.sum(Transaction.amount).label('total')
        ).filter(
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
            Transaction.amount < 0
        ).group_by(
            extract('year', Transaction.transaction_date),
            extract('month', Transaction.transaction_date)
        ).order_by('year', 'month').all()
        
        # 데이터 정리
        income_dict = {(int(row.year), int(row.month)): float(row.total or 0) for row in monthly_income_raw}
        expense_dict = {(int(row.year), int(row.month)): float(row.total or 0) for row in monthly_expense_raw}
        
        all_months = set(income_dict.keys()) | set(expense_dict.keys())
        monthly_data = []
        for year, month in sorted(all_months):
            income = income_dict.get((year, month), 0)
            expense = expense_dict.get((year, month), 0)
            monthly_data.append({
                'period': f'{year}-{month:02d}',
                'income': income,
                'expense': abs(expense),
                'net': income + expense
            })
        
        # 파일 생성
        if export_format == 'csv':
            return export_csv(monthly_data, report_type, start_date, end_date)
        elif export_format == 'excel':
            return export_excel(monthly_data, report_type, start_date, end_date)
        elif export_format == 'pdf':
            return export_pdf(monthly_data, report_type, start_date, end_date)
        else:
            flash('지원하지 않는 형식입니다.', 'error')
            return redirect(url_for('reports'))
            
    except Exception as e:
        flash(f'내보내기 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('reports'))

def export_csv(data, report_type, start_date, end_date):
    """CSV 내보내기"""
    import csv
    import io
    from flask import make_response
    
    output = io.StringIO()
    
    # BOM 추가로 Excel에서 한글 깨짐 방지
    output.write('\ufeff')
    
    writer = csv.writer(output)
    
    # 리포트 헤더 정보
    writer.writerow(['Vlan24 - 재무 리포트'])
    writer.writerow([f'기간: {start_date} ~ {end_date}'])
    writer.writerow([f'생성일: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
    writer.writerow([''])
    
    # 요약 정보
    total_income = sum(row['income'] for row in data)
    total_expense = sum(row['expense'] for row in data)
    net_flow = total_income - total_expense
    
    writer.writerow(['== 요약 정보 =='])
    writer.writerow(['총 수입', f'{total_income:,.0f}원'])
    writer.writerow(['총 지출', f'{total_expense:,.0f}원'])
    writer.writerow(['순현금흐름', f'{net_flow:,.0f}원'])
    writer.writerow(['평균 월 수입', f'{total_income/len(data) if data else 0:,.0f}원'])
    writer.writerow(['평균 월 지출', f'{total_expense/len(data) if data else 0:,.0f}원'])
    writer.writerow([''])
    
    # 월별 상세 데이터
    writer.writerow(['== 월별 상세 =='])
    writer.writerow(['기간', '수입(원)', '지출(원)', '순현금흐름(원)', '수입 비중(%)', '지출 비중(%)'])
    
    for row in data:
        income_ratio = (row['income'] / total_income * 100) if total_income > 0 else 0
        expense_ratio = (row['expense'] / total_expense * 100) if total_expense > 0 else 0
        
        writer.writerow([
            row['period'],
            f"{row['income']:,.0f}",
            f"{row['expense']:,.0f}",
            f"{row['net']:,.0f}",
            f"{income_ratio:.1f}%",
            f"{expense_ratio:.1f}%"
        ])
    
    # 응답 생성 - UTF-8 바이트로 변환
    csv_content = output.getvalue().encode('utf-8-sig')
    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_{end_date}.csv"'
    
    return response

def export_excel(data, report_type, start_date, end_date):
    """Excel 내보내기"""
    import pandas as pd
    import io
    from flask import make_response
    from datetime import datetime
    
    # 요약 정보 계산
    total_income = sum(row['income'] for row in data)
    total_expense = sum(row['expense'] for row in data)
    net_flow = total_income - total_expense
    avg_income = total_income / len(data) if data else 0
    avg_expense = total_expense / len(data) if data else 0
    
    # 요약 데이터프레임
    summary_data = {
        '항목': ['총 수입', '총 지출', '순현금흐름', '평균 월 수입', '평균 월 지출', '분석 기간'],
        '금액': [
            f'{total_income:,.0f}원',
            f'{total_expense:,.0f}원', 
            f'{net_flow:,.0f}원',
            f'{avg_income:,.0f}원',
            f'{avg_expense:,.0f}원',
            f'{len(data)}개월'
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    
    # 상세 데이터프레임
    detail_data = []
    for row in data:
        income_ratio = (row['income'] / total_income * 100) if total_income > 0 else 0
        expense_ratio = (row['expense'] / total_expense * 100) if total_expense > 0 else 0
        
        detail_data.append({
            '기간': row['period'],
            '수입(원)': row['income'],
            '지출(원)': row['expense'],
            '순현금흐름(원)': row['net'],
            '수입 비중(%)': round(income_ratio, 1),
            '지출 비중(%)': round(expense_ratio, 1),
            '수익률(%)': round((row['net'] / row['expense'] * 100) if row['expense'] > 0 else 0, 1)
        })
    
    detail_df = pd.DataFrame(detail_data)
    
    # 메타데이터
    metadata = {
        '정보': ['리포트명', '생성일시', '기간', '데이터 건수'],
        '값': [
            'Vlan24 재무리포트',
            datetime.now().strftime('%Y년 %m월 %d일 %H:%M'),
            f'{start_date} ~ {end_date}',
            f'{len(data)}개월'
        ]
    }
    metadata_df = pd.DataFrame(metadata)
    
    # Excel 파일 생성
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 메타데이터 시트
        metadata_df.to_excel(writer, sheet_name='리포트정보', index=False)
        
        # 요약 시트
        summary_df.to_excel(writer, sheet_name='재무요약', index=False)
        
        # 상세 데이터 시트
        detail_df.to_excel(writer, sheet_name='월별상세', index=False)
        
        # 워크시트 서식 설정
        workbook = writer.book
        
        # 요약 시트 서식
        summary_ws = writer.sheets['재무요약']
        summary_ws.column_dimensions['A'].width = 15
        summary_ws.column_dimensions['B'].width = 20
        
        # 상세 시트 서식
        detail_ws = writer.sheets['월별상세']
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            detail_ws.column_dimensions[col].width = 15
    
    output.seek(0)
    
    # 응답 생성
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_{end_date}.xlsx"'
    
    return response

def export_pdf(data, report_type, start_date, end_date):
    """PDF 내보내기 - 현재 화면 기반"""
    import weasyprint
    from flask import make_response
    from datetime import datetime
    
    # 요약 정보 계산
    total_income = sum(row['income'] for row in data)
    total_expense = sum(row['expense'] for row in data)
    net_flow = total_income - total_expense
    avg_income = total_income / len(data) if data else 0
    avg_expense = total_expense / len(data) if data else 0
    
    # HTML 템플릿 - 영어 버전
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Financial Report</title>
        <style>
            @page {{
                size: A4;
                margin: 1.5cm;
            }}
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                background-color: white;
            }}
            .container {{
                max-width: 100%;
                padding: 20px;
            }}
            .header {{
                text-align: center;
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 3px solid #0d6efd;
            }}
            .header h1 {{
                color: #0d6efd;
                font-size: 28px;
                margin-bottom: 10px;
                font-weight: bold;
            }}
            .header p {{
                color: #666;
                font-size: 16px;
                margin: 5px 0;
            }}
            .section {{
                margin: 25px 0;
                page-break-inside: avoid;
            }}
            .section-title {{
                background: linear-gradient(135deg, #198754, #20c997);
                color: white;
                padding: 12px 20px;
                font-size: 18px;
                font-weight: bold;
                margin-bottom: 15px;
                border-radius: 8px;
            }}
            .summary-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin-bottom: 30px;
            }}
            .summary-card {{
                background: #f8f9fa;
                border: 1px solid #e9ecef;
                border-radius: 10px;
                padding: 20px;
                text-align: center;
            }}
            .summary-card h3 {{
                color: #0d6efd;
                font-size: 14px;
                margin-bottom: 10px;
                text-transform: uppercase;
            }}
            .summary-card .value {{
                font-size: 24px;
                font-weight: bold;
                color: #333;
            }}
            .summary-card.positive .value {{
                color: #198754;
            }}
            .summary-card.negative .value {{
                color: #dc3545;
            }}
            .table-wrapper {{
                background: white;
                border-radius: 10px;
                overflow: hidden;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            .data-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 13px;
            }}
            .data-table thead th {{
                background: #198754;
                color: white;
                padding: 15px 10px;
                text-align: center;
                font-weight: bold;
                border: none;
            }}
            .data-table tbody td {{
                padding: 12px 10px;
                text-align: center;
                border-bottom: 1px solid #e9ecef;
            }}
            .data-table tbody tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            .data-table tbody tr:hover {{
                background-color: #e3f2fd;
            }}
            .positive-amount {{
                color: #198754;
                font-weight: bold;
            }}
            .negative-amount {{
                color: #dc3545;
                font-weight: bold;
            }}
            .monthly-summary {{
                background: linear-gradient(135deg, #f8f9fa, #e9ecef);
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                border-left: 5px solid #0d6efd;
            }}
            .monthly-summary h4 {{
                color: #0d6efd;
                margin-bottom: 15px;
                font-size: 16px;
            }}
            .summary-row {{
                display: flex;
                justify-content: space-between;
                margin: 8px 0;
                padding: 5px 0;
                border-bottom: 1px dotted #ccc;
            }}
            .summary-row:last-child {{
                border-bottom: none;
                font-weight: bold;
                font-size: 16px;
                color: #0d6efd;
            }}
            .footer {{
                margin-top: 30px;
                padding-top: 20px;
                border-top: 2px solid #e9ecef;
                text-align: center;
                color: #666;
                font-size: 12px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Korean Open Banking Accounting System</h1>
                <p><strong>Financial Report</strong></p>
                <p>Period: {start_date} ~ {end_date}</p>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            </div>
            
            <div class="section">
                <div class="section-title">💰 Financial Summary</div>
                <div class="summary-grid">
                    <div class="summary-card">
                        <h3>Total Income</h3>
                        <div class="value positive">{total_income:,.0f} KRW</div>
                    </div>
                    <div class="summary-card">
                        <h3>Total Expenses</h3>
                        <div class="value negative">{total_expense:,.0f} KRW</div>
                    </div>
                    <div class="summary-card {'positive' if net_flow >= 0 else 'negative'}">
                        <h3>Net Cash Flow</h3>
                        <div class="value">{net_flow:,.0f} KRW</div>
                    </div>
                    <div class="summary-card">
                        <h3>Analysis Period</h3>
                        <div class="value">{len(data)} months</div>
                    </div>
                </div>
                
                <div class="monthly-summary">
                    <h4>📊 Monthly Average Analysis</h4>
                    <div class="summary-row">
                        <span>Average Monthly Income:</span>
                        <span class="positive-amount">{avg_income:,.0f} KRW</span>
                    </div>
                    <div class="summary-row">
                        <span>Average Monthly Expenses:</span>
                        <span class="negative-amount">{avg_expense:,.0f} KRW</span>
                    </div>
                    <div class="summary-row">
                        <span>Average Monthly Net Profit:</span>
                        <span class="{'positive-amount' if (avg_income - avg_expense) >= 0 else 'negative-amount'}">{avg_income - avg_expense:,.0f} KRW</span>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">📈 Monthly Details</div>
                <div class="table-wrapper">
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Period</th>
                                <th>Income</th>
                                <th>Expenses</th>
                                <th>Net Cash Flow</th>
                                <th>Income Ratio</th>
                                <th>Expense Ratio</th>
                            </tr>
                        </thead>
                        <tbody>
    """
    
    # 상세 데이터 추가
    for row in data:
        income_ratio = (row['income'] / total_income * 100) if total_income > 0 else 0
        expense_ratio = (row['expense'] / total_expense * 100) if total_expense > 0 else 0
        net_class = 'positive-amount' if row['net'] >= 0 else 'negative-amount'
        
        html_content += f"""
                            <tr>
                                <td><strong>{row['period']}</strong></td>
                                <td class="positive-amount">{row['income']:,.0f} KRW</td>
                                <td class="negative-amount">{row['expense']:,.0f} KRW</td>
                                <td class="{net_class}">{row['net']:,.0f} KRW</td>
                                <td>{income_ratio:.1f}%</td>
                                <td>{expense_ratio:.1f}%</td>
                            </tr>
        """
    
    # 분석 의견 생성
    if net_flow > 0:
        analysis_status = "Good"
        analysis_icon = "✅"
        analysis_text = f"During the analysis period, there was a net cash inflow of <strong>{net_flow:,.0f} KRW</strong>."
    else:
        analysis_status = "Caution"
        analysis_icon = "⚠️"
        analysis_text = f"During the analysis period, there was a net cash outflow of <strong>{abs(net_flow):,.0f} KRW</strong>."
    
    # 추가 분석
    best_month = max(data, key=lambda x: x['net'])['period'] if data else 'N/A'
    worst_month = min(data, key=lambda x: x['net'])['period'] if data else 'N/A'
    
    html_content += f"""
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">📋 Financial Analysis Report</div>
                <div class="monthly-summary">
                    <h4>{analysis_icon} Overall Financial Status: {analysis_status}</h4>
                    <div style="margin: 15px 0; font-size: 14px; line-height: 1.8;">
                        {analysis_text}
                    </div>
                    
                    <div style="margin-top: 20px;">
                        <h4>🔍 Key Metrics Analysis</h4>
                        <div class="summary-row">
                            <span>Best Performance Month:</span>
                            <span class="positive-amount">{best_month}</span>
                        </div>
                        <div class="summary-row">
                            <span>Worst Performance Month:</span>
                            <span class="negative-amount">{worst_month}</span>
                        </div>
                        <div class="summary-row">
                            <span>Expense to Income Ratio:</span>
                            <span>{(total_expense/total_income*100) if total_income > 0 else 0:.1f}%</span>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="footer">
                <p>This report was automatically generated by the Korean Open Banking Accounting System.</p>
                <p>Regular review is recommended for data accuracy.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # WeasyPrint로 PDF 생성
    pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
    
    # 응답 생성
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_{end_date}.pdf"'
    
    return response

@app.route('/budgets')
@login_required
def budgets():
    """예산 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import Department, Category, CategoryBudget
    from datetime import datetime
    
    departments = Department.query.all()
    categories = Category.query.all()
    
    # 현재 월의 분류별 예산 가져오기
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    category_budgets = CategoryBudget.query.filter_by(
        year=current_year, 
        month=current_month, 
        is_active=True
    ).all()
    
    return render_template('budgets.html', 
                         departments=departments,
                         categories=categories,
                         category_budgets=category_budgets,
                         current_year=current_year,
                         current_month=current_month)

@app.route('/budgets/update', methods=['POST'])
@login_required
def update_budgets():
    """예산 업데이트"""
    try:
        from models import Department
        
        for dept_id, budget_str in request.form.items():
            if dept_id.startswith('budget_'):
                department_id = int(dept_id.replace('budget_', ''))
                budget_amount = float(budget_str) if budget_str else 0
                
                department = Department.query.get(department_id)
                if department:
                    department.budget = budget_amount
        
        db.session.commit()
        flash('예산이 성공적으로 업데이트되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'예산 업데이트 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('budgets'))

@app.route('/budgets/category/add', methods=['POST'])
@login_required
def add_category_budget():
    """분류별 예산 추가"""
    try:
        from models import CategoryBudget
        from datetime import datetime
        
        category_id = int(request.form.get('category_id'))
        budget_amount = float(request.form.get('budget_amount', 0))
        year = int(request.form.get('year', datetime.now().year))
        month = int(request.form.get('month', datetime.now().month))
        description = request.form.get('description', '').strip()
        
        # 중복 체크
        existing = CategoryBudget.query.filter_by(
            category_id=category_id,
            year=year,
            month=month,
            is_active=True
        ).first()
        
        if existing:
            flash('해당 분류의 예산이 이미 설정되어 있습니다.', 'error')
            return redirect(url_for('budgets'))
        
        category_budget = CategoryBudget(
            category_id=category_id,
            budget_amount=budget_amount,
            year=year,
            month=month,
            description=description
        )
        
        db.session.add(category_budget)
        db.session.commit()
        
        flash('분류별 예산이 추가되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류별 예산 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('budgets'))

@app.route('/budgets/category/<int:budget_id>/edit', methods=['POST'])
@login_required
def edit_category_budget(budget_id):
    """분류별 예산 수정"""
    try:
        from models import CategoryBudget
        
        category_budget = CategoryBudget.query.get_or_404(budget_id)
        
        category_budget.budget_amount = float(request.form.get('budget_amount', 0))
        category_budget.description = request.form.get('description', '').strip()
        
        db.session.commit()
        
        flash('분류별 예산이 수정되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류별 예산 수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('budgets'))

@app.route('/budgets/category/<int:budget_id>/delete', methods=['POST'])
@login_required
def delete_category_budget(budget_id):
    """분류별 예산 삭제"""
    try:
        from models import CategoryBudget
        
        category_budget = CategoryBudget.query.get_or_404(budget_id)
        category_budget.is_active = False  # Soft delete
        
        db.session.commit()
        
        flash('분류별 예산이 삭제되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류별 예산 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('budgets'))

@app.route('/departments')
@login_required
def departments():
    """부서 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import Department, User
    departments = Department.query.order_by(Department.name).all()
    
    # 통계 계산
    total_departments = len(departments)
    total_budget = sum((dept.budget or 0) for dept in departments)
    total_users = User.query.filter(User.department_id.isnot(None)).count()
    avg_budget = (total_budget / total_departments) if total_departments > 0 else 0
    
    stats = {
        'total_departments': total_departments,
        'total_budget': total_budget,
        'total_users': total_users,
        'avg_budget': avg_budget
    }
    
    return render_template('departments.html', departments=departments, stats=stats)

@app.route('/departments/add', methods=['POST'])
@login_required
def add_department():
    """부서 추가"""
    try:
        from models import Department
        
        name = request.form.get('name', '').strip()
        budget_str = request.form.get('budget', '').strip()
        
        if not name:
            flash('부서명은 필수입니다.', 'error')
            return redirect(url_for('departments'))
        
        # 예산 처리
        budget = None
        if budget_str:
            try:
                budget = float(budget_str.replace(',', ''))
                if budget < 0:
                    flash('예산은 0 이상이어야 합니다.', 'error')
                    return redirect(url_for('departments'))
            except ValueError:
                flash('올바른 예산 금액을 입력해주세요.', 'error')
                return redirect(url_for('departments'))
        
        # 중복 체크
        existing = Department.query.filter_by(name=name).first()
        if existing:
            flash('이미 존재하는 부서명입니다.', 'error')
            return redirect(url_for('departments'))
        
        # 부서 코드 자동 생성
        dept_count = Department.query.count()
        code = f"DEPT{dept_count + 1:03d}"  # DEPT001, DEPT002, ...
        
        # 코드 중복 확인 (만약을 위해)
        while Department.query.filter_by(code=code).first():
            dept_count += 1
            code = f"DEPT{dept_count + 1:03d}"
        
        department = Department(name=name, code=code, budget=budget)
        db.session.add(department)
        db.session.commit()
        
        flash(f'부서 "{name}"이 추가되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'부서 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('departments'))

@app.route('/departments/<int:dept_id>/edit', methods=['POST'])
@login_required
def edit_department(dept_id):
    """부서 수정"""
    try:
        from models import Department
        
        department = Department.query.get_or_404(dept_id)
        name = request.form.get('name', '').strip()
        budget_str = request.form.get('budget', '').strip()
        
        if not name:
            flash('부서명은 필수입니다.', 'error')
            return redirect(url_for('departments'))
        
        # 예산 처리
        budget = None
        if budget_str:
            try:
                budget = float(budget_str.replace(',', ''))
                if budget < 0:
                    flash('예산은 0 이상이어야 합니다.', 'error')
                    return redirect(url_for('departments'))
            except ValueError:
                flash('올바른 예산 금액을 입력해주세요.', 'error')
                return redirect(url_for('departments'))
        
        # 중복 체크 (자기 자신 제외)
        existing = Department.query.filter(
            Department.name == name, 
            Department.id != dept_id
        ).first()
        if existing:
            flash('이미 존재하는 부서명입니다.', 'error')
            return redirect(url_for('departments'))
        
        department.name = name
        department.budget = budget
        # 부서 코드는 수정하지 않고 기존 값 유지
        db.session.commit()
        
        flash(f'부서 "{name}"이 수정되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'부서 수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('departments'))

@app.route('/departments/<int:dept_id>/delete', methods=['POST'])
@login_required
def delete_department(dept_id):
    """부서 삭제"""
    try:
        from models import Department, User
        
        department = Department.query.get_or_404(dept_id)
        
        # 부서에 속한 사용자들의 부서를 None으로 설정
        users_in_department = User.query.filter_by(department_id=dept_id).all()
        for user in users_in_department:
            user.department_id = None
        
        db.session.delete(department)
        db.session.commit()
        
        users_count = len(users_in_department)
        if users_count > 0:
            flash(f'부서 "{department.name}"이 삭제되었습니다. {users_count}명의 직원이 부서 없음 상태로 변경되었습니다.', 'success')
        else:
            flash(f'부서 "{department.name}"이 삭제되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'부서 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('departments'))


@app.route('/categories')
@login_required
def categories():
    """분류 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import Category
    categories = Category.query.all()
    return render_template('categories.html', categories=categories)


@app.route('/categories/add', methods=['POST'])
@login_required
def add_category():
    """분류 추가"""
    try:
        from models import Category
        
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            flash('분류명은 필수입니다.', 'error')
            return redirect(url_for('categories'))
        
        # 분류명 중복 체크
        existing_name = Category.query.filter_by(name=name).first()
        if existing_name:
            flash('이미 존재하는 분류명입니다.', 'error')
            return redirect(url_for('categories'))
        
        # 자동 코드 생성
        last_category = Category.query.order_by(Category.id.desc()).first()
        if last_category:
            # 마지막 분류의 ID를 기반으로 다음 코드 생성
            next_id = last_category.id + 1
        else:
            next_id = 1
        
        code = f"CAT{next_id:03d}"
        
        # 혹시 코드 중복이 있는지 확인 (안전장치)
        while Category.query.filter_by(code=code).first():
            next_id += 1
            code = f"CAT{next_id:03d}"
        
        category = Category(code=code, name=name, description=description)
        db.session.add(category)
        db.session.commit()
        
        flash(f'분류 "{name}" (코드: {code})이 추가되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('categories'))


@app.route('/categories/<int:category_id>/edit', methods=['POST'])
@login_required
def edit_category(category_id):
    """분류 수정"""
    try:
        from models import Category
        
        category = Category.query.get_or_404(category_id)
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            flash('분류명은 필수입니다.', 'error')
            return redirect(url_for('categories'))
        
        # 분류명 중복 체크 (자기 자신 제외)
        existing_name = Category.query.filter(
            Category.name == name, 
            Category.id != category_id
        ).first()
        if existing_name:
            flash('이미 존재하는 분류명입니다.', 'error')
            return redirect(url_for('categories'))
        
        # 코드는 수정하지 않고 이름과 설명만 수정
        category.name = name
        category.description = description
        db.session.commit()
        
        flash(f'분류 "{name}"이 수정되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류 수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('categories'))


@app.route('/categories/<int:category_id>/delete', methods=['POST'])
@login_required
def delete_category(category_id):
    """분류 삭제"""
    try:
        from models import Category, Vendor, Transaction
        
        category = Category.query.get_or_404(category_id)
        
        # 분류에 연결된 업체가 있는지 확인
        vendors_count = Vendor.query.filter_by(category_id=category_id).count()
        if vendors_count > 0:
            flash(f'분류에 {vendors_count}개의 업체가 연결되어 있어 삭제할 수 없습니다.', 'error')
            return redirect(url_for('categories'))
        
        # 분류에 연결된 거래가 있는지 확인
        transactions_count = Transaction.query.filter_by(category_id=category_id).count()
        if transactions_count > 0:
            flash(f'분류에 {transactions_count}개의 거래가 연결되어 있어 삭제할 수 없습니다.', 'error')
            return redirect(url_for('categories'))
        
        db.session.delete(category)
        db.session.commit()
        
        flash(f'분류 "{category.name}"이 삭제되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'분류 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('categories'))


@app.route('/vendors')
@login_required
def vendors():
    """업체 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    from models import Vendor, Category
    vendors = Vendor.query.all()
    categories = Category.query.all()
    return render_template('vendors.html', vendors=vendors, categories=categories)


@app.route('/vendors/add', methods=['POST'])
@login_required
def add_vendor():
    """공급업체 추가"""
    try:
        from models import Vendor
        
        name = request.form.get('name', '').strip()
        business_number = request.form.get('business_number', '').strip()
        contact_info = request.form.get('contact_info', '').strip()
        category_id = request.form.get('category_id')
        
        if not name:
            flash('업체명은 필수입니다.', 'error')
            return redirect(url_for('vendors'))
        
        # 중복 체크
        existing = Vendor.query.filter_by(name=name).first()
        if existing:
            flash('이미 존재하는 업체명입니다.', 'error')
            return redirect(url_for('vendors'))
        
        vendor = Vendor(
            name=name,
            business_number=business_number if business_number else None,
            contact_info=contact_info if contact_info else None,
            category_id=int(category_id) if category_id else None
        )
        
        db.session.add(vendor)
        db.session.commit()
        
        flash(f'공급업체 "{name}"이 추가되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'공급업체 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('vendors'))


@app.route('/vendors/<int:vendor_id>/edit', methods=['POST'])
@login_required
def edit_vendor(vendor_id):
    """업체 수정"""
    try:
        from models import Vendor
        
        vendor = Vendor.query.get_or_404(vendor_id)
        name = request.form.get('name', '').strip()
        business_number = request.form.get('business_number', '').strip()
        contact_info = request.form.get('contact_info', '').strip()
        category_id = request.form.get('category_id')
        
        if not name:
            flash('업체명은 필수입니다.', 'error')
            return redirect(url_for('vendors'))
        
        # 중복 체크 (자기 자신 제외)
        existing = Vendor.query.filter(
            Vendor.name == name, 
            Vendor.id != vendor_id
        ).first()
        if existing:
            flash('이미 존재하는 업체명입니다.', 'error')
            return redirect(url_for('vendors'))
        
        vendor.name = name
        vendor.business_number = business_number if business_number else None
        vendor.contact_info = contact_info if contact_info else None
        vendor.category_id = int(category_id) if category_id else None
        db.session.commit()
        
        flash(f'업체 "{name}"이 수정되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'업체 수정 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('vendors'))


@app.route('/vendors/<int:vendor_id>/delete', methods=['POST'])
@login_required
def delete_vendor(vendor_id):
    """업체 삭제"""
    try:
        from models import Vendor, Contract, Transaction
        
        vendor = Vendor.query.get_or_404(vendor_id)
        
        # 업체와 연결된 계약이 있는지 확인
        contracts_count = Contract.query.filter_by(vendor_id=vendor_id).count()
        if contracts_count > 0:
            flash(f'업체에 {contracts_count}개의 계약이 있어 삭제할 수 없습니다.', 'error')
            return redirect(url_for('vendors'))
        
        # 업체와 연결된 거래가 있는지 확인
        transactions_count = Transaction.query.filter_by(vendor_id=vendor_id).count()
        if transactions_count > 0:
            flash(f'업체에 {transactions_count}개의 거래가 있어 삭제할 수 없습니다.', 'error')
            return redirect(url_for('vendors'))
        
        vendor_name = vendor.name
        db.session.delete(vendor)
        db.session.commit()
        
        flash(f'업체 "{vendor_name}"이 삭제되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'업체 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('vendors'))


# API 엔드포인트들

@app.route('/users')
@login_required
def users():
    """사용자 관리 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    departments = Department.query.all()
    return render_template('users.html', users=users, departments=departments)

@app.route('/user/<int:user_id>/toggle', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    """사용자 활성화/비활성화 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    # 자기 자신은 비활성화할 수 없음
    if user.id == current_user.id:
        flash('자기 자신의 계정은 비활성화할 수 없습니다.', 'error')
        return redirect(url_for('users'))
    
    user.active = not user.active
    db.session.commit()
    
    status = "활성화" if user.active else "비활성화"
    flash(f'{user.name} 사용자가 {status}되었습니다.', 'success')
    return redirect(url_for('users'))

@app.route('/user/<int:user_id>/role', methods=['POST'])
@login_required
def change_user_role(user_id):
    """사용자 권한 변경 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    new_role = request.form.get('role')
    
    if new_role not in ['admin', 'user']:
        flash('잘못된 권한입니다.', 'error')
        return redirect(url_for('users'))
    
    # 자기 자신의 관리자 권한은 해제할 수 없음
    if user.id == current_user.id and new_role != 'admin':
        flash('자기 자신의 관리자 권한은 해제할 수 없습니다.', 'error')
        return redirect(url_for('users'))
    
    old_role = user.role
    user.role = new_role
    db.session.commit()
    
    role_name = "관리자" if new_role == 'admin' else "일반 사용자"
    flash(f'{user.name} 사용자의 권한이 {role_name}로 변경되었습니다.', 'success')
    return redirect(url_for('users'))

@app.route('/user/<int:user_id>/department', methods=['POST'])
@login_required
def change_user_department(user_id):
    """사용자 부서 변경 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    department_id = request.form.get('department_id') or None
    
    user.department_id = department_id
    db.session.commit()
    
    if department_id:
        department = Department.query.get(department_id)
        if department:
            flash(f'{user.name} 사용자의 부서가 {department.name}로 변경되었습니다.', 'success')
        else:
            flash(f'{user.name} 사용자의 부서가 변경되었습니다.', 'success')
    else:
        flash(f'{user.name} 사용자의 부서가 제거되었습니다.', 'success')
    
    return redirect(url_for('users'))

@app.route('/user/add', methods=['POST'])
@login_required
def add_user():
    """새 사용자 추가 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    # 이메일 중복 확인
    email = request.form.get('email')
    if User.query.filter_by(email=email).first():
        flash('이미 존재하는 이메일입니다.', 'error')
        return redirect(url_for('users'))
    
    # 새 사용자 생성
    user = User()
    user.name = request.form.get('name')
    user.email = email
    user.set_password(request.form.get('password'))
    user.role = request.form.get('role', 'user')
    user.department_id = request.form.get('department_id') or None
    
    db.session.add(user)
    db.session.commit()
    
    flash(f'{user.name} 사용자가 추가되었습니다.', 'success')
    return redirect(url_for('users'))

@app.route('/user/<int:user_id>/edit', methods=['POST'])
@login_required
def edit_user(user_id):
    """사용자 정보 수정 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    # 폼 데이터 가져오기
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    current_password = request.form.get('current_password', '').strip()
    password = request.form.get('password', '').strip()
    password_confirm = request.form.get('password_confirm', '').strip()
    role = request.form.get('role')
    department_id = request.form.get('department_id') or None
    
    # 이름 검증
    if not name:
        flash('이름은 필수입니다.', 'error')
        return redirect(url_for('users'))
    
    # 이메일 검증 및 중복 확인
    if not email:
        flash('이메일은 필수입니다.', 'error')
        return redirect(url_for('users'))
    
    existing_user = User.query.filter_by(email=email).first()
    if existing_user and existing_user.id != user.id:
        flash('이미 존재하는 이메일입니다.', 'error')
        return redirect(url_for('users'))
    
    # 권한 검증
    if role not in ['admin', 'user']:
        flash('잘못된 권한입니다.', 'error')
        return redirect(url_for('users'))
    
    # 자기 자신의 관리자 권한 해제 방지
    if user.id == current_user.id and role != 'admin':
        flash('자기 자신의 관리자 권한은 해제할 수 없습니다.', 'error')
        return redirect(url_for('users'))
    
    # 정보 업데이트
    user.name = name
    user.email = email
    user.role = role
    user.department_id = department_id
    
    # 패스워드 업데이트 (비어있지 않은 경우만)
    if password:
        # 패스워드 길이 검증
        if len(password) < 6:
            flash('패스워드는 최소 6자 이상이어야 합니다.', 'error')
            return redirect(url_for('users'))
        
        # 패스워드 확인 검증
        if password != password_confirm:
            flash('새 패스워드가 일치하지 않습니다.', 'error')
            return redirect(url_for('users'))
        
        # 자기 자신의 패스워드를 변경하는 경우 현재 패스워드 확인
        if user.id == current_user.id:
            if not current_password:
                flash('자신의 패스워드를 변경하려면 현재 패스워드를 입력해야 합니다.', 'error')
                return redirect(url_for('users'))
            
            if not user.check_password(current_password):
                flash('현재 패스워드가 올바르지 않습니다.', 'error')
                return redirect(url_for('users'))
        
        user.set_password(password)
    
    db.session.commit()
    
    # 감사 로그 기록
    audit_log = AuditLog()
    audit_log.user_id = str(current_user.id)
    audit_log.action = '사용자 정보 수정'
    audit_log.table_name = 'users'
    audit_log.record_id = user.id
    audit_log.new_values = f'{user.name} 사용자 정보 수정 완료'
    audit_log.created_at = datetime.utcnow()
    
    db.session.add(audit_log)
    db.session.commit()
    
    flash(f'{user.name} 사용자 정보가 성공적으로 업데이트되었습니다.', 'success')
    return redirect(url_for('users'))

@app.route('/user/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """사용자 삭제 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    # 자기 자신은 삭제할 수 없음
    if user.id == current_user.id:
        flash('자기 자신의 계정은 삭제할 수 없습니다.', 'error')
        return redirect(url_for('users'))
    
    user_name = user.name
    db.session.delete(user)
    db.session.commit()
    
    flash(f'{user_name} 사용자가 삭제되었습니다.', 'success')
    return redirect(url_for('users'))

@app.route('/api/dashboard/chart-data')
@login_required
def api_dashboard_chart_data():
    """대시보드 차트 데이터 API"""
    # 최근 7일 일별 현금흐름
    end_date = date.today()
    start_date = end_date - timedelta(days=6)
    
    # 간단하게 모든 거래를 가져와서 처리
    transactions = Transaction.query.filter(
        func.date(Transaction.transaction_date) >= start_date
    ).all()
    
    # 날짜별로 데이터 집계
    daily_data = {}
    current_date = start_date
    while current_date <= end_date:
        daily_data[current_date] = {'income': 0, 'expense': 0}
        current_date += timedelta(days=1)
    
    for transaction in transactions:
        trans_date = transaction.transaction_date.date()
        if trans_date in daily_data:
            if transaction.amount > 0:
                daily_data[trans_date]['income'] += float(transaction.amount)
            else:
                daily_data[trans_date]['expense'] += float(abs(transaction.amount))
    
    dates = []
    income_data = []
    expense_data = []
    
    for date_key in sorted(daily_data.keys()):
        dates.append(date_key.strftime('%m/%d'))
        income_data.append(daily_data[date_key]['income'])
        expense_data.append(daily_data[date_key]['expense'])
    
    return jsonify({
        'dates': dates,
        'income': income_data,
        'expense': expense_data
    })

# 초기 데이터 생성용 헬퍼 함수들

def init_sample_data():
    """샘플 데이터 초기화 (개발용)"""
    if Institution.query.count() > 0:
        return
    
    # 금융기관 데이터
    institutions = []
    institution_data = [
        {'code': '001', 'name': 'KB국민은행', 'type': 'bank'},
        {'code': '002', 'name': '신한은행', 'type': 'bank'},
        {'code': '003', 'name': '우리은행', 'type': 'bank'},
        {'code': '101', 'name': '삼성카드', 'type': 'card'},
        {'code': '102', 'name': '현대카드', 'type': 'card'},
    ]
    for data in institution_data:
        inst = Institution()
        inst.code = data['code']
        inst.name = data['name']
        inst.type = data['type']
        institutions.append(inst)
    
    # 부서 데이터
    departments = []
    department_data = [
        {'code': '001', 'name': '경영지원팀', 'budget': 10000000},
        {'code': '002', 'name': '개발팀', 'budget': 15000000},
        {'code': '003', 'name': '마케팅팀', 'budget': 8000000},
        {'code': '004', 'name': '영업팀', 'budget': 12000000},
    ]
    for data in department_data:
        dept = Department()
        dept.code = data['code']
        dept.name = data['name']
        dept.budget = data['budget']
        departments.append(dept)
    
    # 카테고리 데이터
    categories = []
    category_data = [
        {'code': '001', 'name': '사무용품'},
        {'code': '002', 'name': '교통비'},
        {'code': '003', 'name': '식비'},
        {'code': '004', 'name': '임대료'},
        {'code': '005', 'name': '통신비'},
        {'code': '006', 'name': '광고비'},
        {'code': '007', 'name': '회의비'},
    ]
    for data in category_data:
        cat = Category()
        cat.code = data['code']
        cat.name = data['name']
        categories.append(cat)
    
    # 거래처 데이터
    vendors = []
    vendor_data = [
        {'name': '사무용품쇼핑몰', 'business_number': '123-45-67890', 'category_id': 1},
        {'name': '카카오T', 'business_number': '234-56-78901', 'category_id': 2},
        {'name': '배달의민족', 'business_number': '345-67-89012', 'category_id': 3},
        {'name': '부동산관리공사', 'business_number': '456-78-90123', 'category_id': 4},
        {'name': 'SKT', 'business_number': '567-89-01234', 'category_id': 5},
        {'name': '스타벅스', 'business_number': '678-90-12345', 'category_id': 3},
        {'name': '네이버', 'business_number': '789-01-23456', 'category_id': 6},
    ]
    for data in vendor_data:
        vendor = Vendor()
        vendor.name = data['name']
        vendor.business_number = data['business_number']
        vendor.category_id = data['category_id']
        vendors.append(vendor)
    
    db.session.add_all(institutions + departments + categories + vendors)
    db.session.commit()
    
    # 계좌 데이터
    accounts = []
    account_data = [
        {'institution_id': 1, 'account_number': '123-456-789012', 'account_name': '법인통장', 'account_type': 'checking', 'balance': 50000000, 'department_id': 1},
        {'institution_id': 2, 'account_number': '987-654-321098', 'account_name': '개발팀 통장', 'account_type': 'checking', 'balance': 15000000, 'department_id': 2},
        {'institution_id': 4, 'account_number': '1234-5678-9012', 'account_name': '법인카드', 'account_type': 'credit', 'balance': 0, 'department_id': 1},
    ]
    for data in account_data:
        account = Account()
        account.institution_id = data['institution_id']
        account.account_number = data['account_number']
        account.account_name = data['account_name']
        account.account_type = data['account_type']
        account.balance = data['balance']
        account.department_id = data['department_id']
        accounts.append(account)
    
    db.session.add_all(accounts)
    db.session.commit()
    
    # 샘플 거래 데이터
    sample_transactions = []
    for i in range(100):
        transaction = Transaction()
        transaction.account_id = 1
        transaction.transaction_id = f'TXN-{i:06d}'
        transaction.amount = -5500 if i % 3 == 0 else (15000 if i % 5 == 0 else -12000)
        transaction.transaction_type = 'debit'
        transaction.description = f'스타벅스 강남점' if i % 3 == 0 else (f'프로젝트 수수료 입금' if i % 5 == 0 else f'사무용품 구매')
        transaction.counterparty = f'스타벅스' if i % 3 == 0 else (f'클라이언트' if i % 5 == 0 else f'오피스디포')
        transaction.transaction_date = datetime.now() - timedelta(days=i)
        transaction.classification_status = 'pending' if i % 4 == 0 else 'classified'
        transaction.category_id = 3 if i % 3 == 0 else (None if i % 5 == 0 else 1)
        transaction.department_id = 2 if i % 2 == 0 else 1
        transaction.vendor_id = 6 if i % 3 == 0 else (None if i % 5 == 0 else 1)
        sample_transactions.append(transaction)
    
    db.session.add_all(sample_transactions)
    db.session.commit()
    
    # 샘플 알림 데이터
    sample_alerts = []
    alert_data = [
        {'title': '예산 초과 경고', 'message': '개발팀의 이번 달 지출이 예산의 85%에 달했습니다.', 'alert_type': 'budget', 'severity': 'warning'},
        {'title': '계약 만료 임박', 'message': 'SKT 통신 서비스 계약이 30일 후 만료됩니다.', 'alert_type': 'contract', 'severity': 'info'},
        {'title': '이상거래 감지', 'message': '평소보다 큰 금액의 거래가 감지되었습니다. (1,500,000원)', 'alert_type': 'anomaly', 'severity': 'warning'},
    ]
    for data in alert_data:
        alert = Alert()
        alert.title = data['title']
        alert.message = data['message']
        alert.alert_type = data['alert_type']
        alert.severity = data['severity']
        sample_alerts.append(alert)
    
    db.session.add_all(sample_alerts)
    db.session.commit()
    
    # 샘플 분류 규칙
    sample_rules = []
    rule_data = [
        {'name': '스타벅스 자동분류', 'priority': 8, 'condition_type': 'contains', 'condition_field': 'counterparty', 'condition_value': '스타벅스', 'target_category_id': 3, 'target_vendor_id': 6},
        {'name': '사무용품 자동분류', 'priority': 7, 'condition_type': 'contains', 'condition_field': 'description', 'condition_value': '사무용품', 'target_category_id': 1, 'target_department_id': 1, 'target_vendor_id': 1},
    ]
    for data in rule_data:
        rule = MappingRule()
        rule.name = data['name']
        rule.priority = data['priority']
        rule.condition_type = data['condition_type']
        rule.condition_field = data['condition_field']
        rule.condition_value = data['condition_value']
        rule.target_category_id = data.get('target_category_id')
        rule.target_department_id = data.get('target_department_id')
        rule.target_vendor_id = data.get('target_vendor_id')
        sample_rules.append(rule)
    
    db.session.add_all(sample_rules)
    db.session.commit()

# 데이터 관리 라우트
@app.route('/data-management')
@login_required
def data_management():
    """데이터 관리 페이지 (관리자 전용)"""
    if not current_user.is_admin():
        flash('관리자만 접근할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    accounts = Account.query.all()
    categories = Category.query.all()
    departments = Department.query.all()
    vendors = Vendor.query.all()
    # 최근 업로드 기록 조회 (활성 거래만)
    recent_uploads_query = db.session.query(
        func.date(Transaction.created_at).label('upload_date'),
        func.max(Transaction.created_at).label('latest_time'),
        Transaction.account_id,
        func.count(Transaction.id).label('processed_count'),
        func.max(Transaction.id).label('max_id')
    ).filter(
        Transaction.is_active == True
    ).group_by(
        func.date(Transaction.created_at),
        Transaction.account_id
    ).order_by(func.max(Transaction.created_at).desc()).limit(10)
    
    recent_uploads = []
    for upload in recent_uploads_query:
        account = Account.query.get(upload.account_id)
        recent_uploads.append({
            'id': upload.max_id,
            'created_at': upload.latest_time,
            'filename': f"transactions_{upload.upload_date}.csv",
            'account': account,
            'processed_count': upload.processed_count,
            'status': 'completed'
        })
    
    return render_template('data_management.html', 
                         accounts=accounts,
                         categories=categories,
                         departments=departments,
                         vendors=vendors,
                         recent_uploads=recent_uploads)


@app.route('/export-transactions-period', methods=['POST'])
@login_required
def export_transactions_period():
    """기간별 거래 내역 다운로드"""
    try:
        import io
        from datetime import datetime
        from models import Transaction, Account, Category, Department, Vendor
        
        # 폼 데이터 받기
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        export_format = request.form.get('export_format', 'csv')
        
        if not start_date_str or not end_date_str:
            flash('시작일과 종료일을 모두 입력해주세요.', 'error')
            return redirect(url_for('data_management'))
        
        # 날짜 파싱
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        if start_date > end_date:
            flash('시작일은 종료일보다 이전이어야 합니다.', 'error')
            return redirect(url_for('data_management'))
        
        # 기간내 거래 조회
        transactions = Transaction.query.filter(
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        ).order_by(Transaction.transaction_date.desc()).all()
        
        if not transactions:
            flash(f'{start_date_str}부터 {end_date_str}까지 거래 내역이 없습니다.', 'warning')
            return redirect(url_for('data_management'))
        
        # 데이터 준비
        transaction_data = []
        for transaction in transactions:
            # 관련 데이터 조회
            account = Account.query.get(transaction.account_id) if transaction.account_id else None
            category = Category.query.get(transaction.category_id) if transaction.category_id else None
            department = Department.query.get(transaction.department_id) if transaction.department_id else None
            vendor = Vendor.query.get(transaction.vendor_id) if transaction.vendor_id else None
            
            transaction_data.append({
                '거래일': transaction.transaction_date.strftime('%Y-%m-%d'),
                '거래시간': transaction.transaction_date.strftime('%H:%M:%S') if transaction.transaction_date else '',
                '계정': account.name if account else '',
                '거래유형': transaction.transaction_type or '',
                '금액': transaction.amount,
                '잔액': transaction.balance if transaction.balance else '',
                '내용': transaction.description or '',
                '상대계좌': transaction.counterpart_account or '',
                '상대은행': transaction.counterpart_bank or '',
                '분류': category.name if category else '',
                '부서': department.name if department else '',
                '업체': vendor.name if vendor else '',
                '분류상태': {
                    'pending': '미분류',
                    'classified': '분류완료',
                    'manual': '수동분류'
                }.get(transaction.classification_status, transaction.classification_status or ''),
                '메모': transaction.memo or ''
            })
        
        # 파일명 생성
        filename = f"transactions_{start_date_str}_{end_date_str}"
        
        if export_format == 'excel':
            # Excel 파일 생성
            import pandas as pd
            df = pd.DataFrame(transaction_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='거래내역', index=False)
            
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f"{filename}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        else:
            # CSV 파일 생성 (UTF-8 BOM으로 한글 지원)
            import csv
            output = io.StringIO()
            
            if transaction_data:
                fieldnames = transaction_data[0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(transaction_data)
            
            # UTF-8 BOM 추가 (Excel에서 한글 깨짐 방지)
            csv_content = '\ufeff' + output.getvalue()
            
            return Response(
                csv_content,
                mimetype='text/csv; charset=utf-8',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}.csv"',
                    'Content-Type': 'text/csv; charset=utf-8'
                }
            )
            
    except Exception as e:
        flash(f'다운로드 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('data_management'))

@app.route('/upload-transactions', methods=['POST'])
@login_required
def upload_transactions():
    """파일 업로드 처리"""
    try:
        if 'transaction_file' not in request.files:
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'})
        
        file = request.files['transaction_file']
        account_id = request.form.get('account_id')
        
        # 새로운 폼 필드들
        default_transaction_type = request.form.get('default_transaction_type', '')
        default_target_account = request.form.get('default_target_account', '')
        default_category_id = request.form.get('default_category_id', '')
        default_department_id = request.form.get('default_department_id', '')
        default_vendor_id = request.form.get('default_vendor_id', '')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'})
        
        if not account_id:
            return jsonify({'success': False, 'error': '계정을 선택해주세요.'})
        
        # 계정 확인
        account = Account.query.get(account_id)
        if not account:
            return jsonify({'success': False, 'error': '유효하지 않은 계정입니다.'})
        
        # 파일 확장자 확인
        if not file.filename:
            return jsonify({'success': False, 'error': '파일명이 유효하지 않습니다.'})
            
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        if file_ext not in ['csv', 'xls', 'xlsx']:
            return jsonify({'success': False, 'error': '지원하지 않는 파일 형식입니다.'})
        
        # 파일 읽기
        if file_ext == 'csv':
            df = pd.read_csv(file.stream, encoding='utf-8')
        else:
            df = pd.read_excel(file.stream)
        
        # 필수 컬럼 확인 (항상 5개 필수)
        required_columns = ['계정', '거래일자', '거래유형', '금액', '거래처']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({'success': False, 'error': f'필수 컬럼이 누락되었습니다: {", ".join(missing_columns)}'})
        
        # 데이터 처리 및 저장
        processed_count = 0
        for _, row in df.iterrows():
            try:
                transaction = Transaction()
                
                # 계정 처리 - 파일에 계정 컬럼이 있으면 그것을 우선 사용
                if '계정' in df.columns and pd.notna(row['계정']):
                    account_name = str(row['계정']).strip()
                    # 계정명으로 계정 찾기
                    file_account = Account.query.filter_by(name=account_name).first()
                    if not file_account:
                        print(f"Warning: Account '{account_name}' not found, using default account")
                        transaction.account_id = account_id  # 기본 계정 사용
                    else:
                        transaction.account_id = file_account.id
                else:
                    # 파일에 계정 컬럼이 없으면 폼에서 선택한 계정 사용
                    transaction.account_id = account_id
                
                transaction.transaction_id = f'UPLOAD-{datetime.now().strftime("%Y%m%d%H%M%S")}-{processed_count:04d}'
                
                # 거래일자 처리
                transaction_date = pd.to_datetime(row['거래일자'])
                if '거래시간' in df.columns and pd.notna(row['거래시간']):
                    # 거래시간이 있으면 결합
                    time_str = str(row['거래시간'])
                    if ':' in time_str:
                        date_str = transaction_date.strftime('%Y-%m-%d')
                        transaction_datetime = pd.to_datetime(f"{date_str} {time_str}")
                        transaction.transaction_date = transaction_datetime
                    else:
                        transaction.transaction_date = transaction_date
                else:
                    transaction.transaction_date = transaction_date
                
                # 금액 처리
                amount = float(str(row['금액']).replace(',', '').replace('원', ''))
                transaction.amount = amount
                
                # 거래유형 처리
                transaction_type = str(row['거래유형']).strip()
                
                if transaction_type in ['입금', '수입', 'deposit']:
                    transaction.transaction_type = 'credit'
                elif transaction_type in ['출금', '지출', 'withdrawal']:
                    transaction.transaction_type = 'debit'
                    transaction.amount = -abs(amount)  # 지출은 음수로
                elif transaction_type in ['이체', 'transfer']:
                    transaction.transaction_type = 'transfer'
                    transaction.amount = -abs(amount)  # 이체도 음수 (보내는 쪽)
                    
                    # 이체의 경우 대상계정 필수 체크
                    if not ('대상계정' in df.columns and pd.notna(row['대상계정']) and str(row['대상계정']).strip()):
                        print(f"Warning: Transfer transaction missing target account in row {processed_count}")
                        continue  # 이체인데 대상계정이 없으면 건너뛰기
                else:
                    transaction.transaction_type = 'debit'  # 기본값
                
                # 거래처 정보
                transaction.counterparty = str(row['거래처']).strip()
                
                # 메모 (선택사항)
                if '메모' in df.columns and pd.notna(row['메모']):
                    transaction.description = str(row['메모']).strip()
                else:
                    transaction.description = transaction.counterparty
                
                # 대상 계정 처리 (이체의 경우)
                if transaction.transaction_type == 'transfer':
                    if '대상계정' in df.columns and pd.notna(row['대상계정']):
                        target_account_name = str(row['대상계정']).strip()
                        transaction.description += f' (받는 계정: {target_account_name})'
                    elif default_target_account:
                        target_account = Account.query.get(default_target_account)
                        if target_account:
                            transaction.description += f' (받는 계정: {target_account.name})'
                
                # 분류 정보 처리 (이체가 아닌 경우만)
                if transaction.transaction_type != 'transfer':
                    # 카테고리
                    if '분류' in df.columns and pd.notna(row['분류']):
                        category_name = str(row['분류']).strip()
                        category = Category.query.filter_by(name=category_name).first()
                        if category:
                            transaction.category_id = category.id
                    elif default_category_id:
                        transaction.category_id = default_category_id
                    
                    # 부서
                    if '부서' in df.columns and pd.notna(row['부서']):
                        dept_name = str(row['부서']).strip()
                        department = Department.query.filter_by(name=dept_name).first()
                        if department:
                            transaction.department_id = department.id
                    elif default_department_id:
                        transaction.department_id = default_department_id
                    
                    # 업체
                    if '업체' in df.columns and pd.notna(row['업체']):
                        vendor_name = str(row['업체']).strip()
                        vendor = Vendor.query.filter_by(name=vendor_name).first()
                        if vendor:
                            transaction.vendor_id = vendor.id
                    elif default_vendor_id:
                        transaction.vendor_id = default_vendor_id
                    
                    # 분류 상태 설정
                    if transaction.category_id or transaction.department_id or transaction.vendor_id:
                        transaction.classification_status = 'classified'
                    else:
                        transaction.classification_status = 'pending'
                else:
                    # 이체는 자동으로 분류됨
                    transaction.classification_status = 'classified'
                
                db.session.add(transaction)
                processed_count += 1
                
            except Exception as e:
                print(f"Row processing error: {e}")
                continue
        
        db.session.commit()
        
        # 자동 분류 규칙 적용 (새로 추가된 거래들에 대해)
        new_transactions = Transaction.query.filter(
            Transaction.transaction_id.like(f'UPLOAD-{datetime.now().strftime("%Y%m%d%H%M%S")}%')
        ).all()
        
        for transaction in new_transactions:
            apply_classification_rules(transaction)
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'processed_count': processed_count,
            'message': f'{processed_count}건의 거래가 성공적으로 업로드되었습니다.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'파일 처리 중 오류가 발생했습니다: {str(e)}'})

def check_and_create_anomaly_alerts(transaction):
    """거래에 대해 이상거래 알림 확인 및 생성"""
    try:
        from models import AlertSetting, Alert
        
        # 활성화된 이상거래 알림 설정 조회
        alert_settings = AlertSetting.query.filter(
            AlertSetting.is_active == True,
            AlertSetting.alert_type == 'anomaly'
        ).all()
        
        for setting in alert_settings:
            should_alert = False
            alert_message = ""
            
            # 금액 기반 조건 확인
            if setting.condition_type == 'amount_range' and setting.condition_field == 'amount':
                # condition_value 형태: "10000,9999999999" (최소값,최대값)
                try:
                    min_amount, max_amount = setting.condition_value.split(',')
                    min_amount = float(min_amount)
                    max_amount = float(max_amount)
                    
                    transaction_amount = abs(transaction.amount)  # 절댓값으로 비교
                    
                    if min_amount <= transaction_amount <= max_amount:
                        should_alert = True
                        alert_message = f"이상 거래 감지: {transaction.counterparty}에서 {abs(transaction.amount):,.0f}원 거래가 발생했습니다."
                        
                except (ValueError, AttributeError):
                    continue
            
            # contains 조건 확인
            elif setting.condition_type == 'contains':
                field_value = ""
                if setting.condition_field == 'description':
                    field_value = transaction.description or ""
                elif setting.condition_field == 'counterparty':
                    field_value = transaction.counterparty or ""
                
                if setting.condition_value.lower() in field_value.lower():
                    should_alert = True
                    alert_message = f"알림 조건 일치: {transaction.counterparty}에서 '{setting.condition_value}' 포함된 거래가 발생했습니다."
            
            # 알림 생성
            if should_alert:
                alert = Alert()
                alert.title = f"[{setting.name}] 이상거래 감지"
                alert.message = alert_message
                alert.alert_type = 'anomaly'
                alert.severity = setting.severity or 'warning'
                alert.related_table = 'transaction'
                alert.related_id = transaction.id
                alert.is_read = False
                
                db.session.add(alert)
                print(f"Anomaly alert created: {alert.title} - {alert.message}")
        
        db.session.commit()
        
    except Exception as e:
        print(f"Error checking anomaly alerts: {str(e)}")
        # 알림 생성 실패해도 거래 추가는 계속 진행


@app.route('/add-transaction', methods=['POST'])
@login_required
def add_transaction():
    """거래 추가"""
    try:
        # 폼 데이터 받기
        account_id = request.form.get('account_id')
        transaction_type = request.form.get('transaction_type')
        transaction_date = request.form.get('transaction_date')
        amount = request.form.get('amount')
        counterparty = request.form.get('counterparty')
        description = request.form.get('description', '')
        target_account_id = request.form.get('target_account_id', '')
        
        # 분류 정보 (이체가 아닌 경우만)
        category_id = request.form.get('category_id') if transaction_type != 'transfer' else None
        department_id = request.form.get('department_id') if transaction_type != 'transfer' else None
        vendor_id = request.form.get('vendor_id') if transaction_type != 'transfer' else None
        
        # 유효성 검사 (거래처는 선택사항)
        if not all([account_id, transaction_type, transaction_date, amount]):
            flash('필수 항목을 모두 입력해주세요.', 'error')
            return redirect(url_for('transactions'))
        
        # 거래처가 비어있으면 기본값 설정
        if not counterparty:
            if transaction_type == 'deposit':
                counterparty = '입금'
            elif transaction_type == 'withdrawal':
                counterparty = '출금'
            elif transaction_type == 'transfer':
                counterparty = '계좌이체'
            else:
                counterparty = '거래처 미분류'
        
        # 계정 확인
        account = Account.query.get(account_id)
        if not account:
            flash('유효하지 않은 계정입니다.', 'error')
            return redirect(url_for('transactions'))
        
        # 새 거래 생성
        transaction = Transaction()
        transaction.account_id = account_id
        transaction.transaction_id = f'MANUAL-{datetime.now().strftime("%Y%m%d%H%M%S")}-{int(datetime.now().timestamp() * 1000) % 10000:04d}'
        transaction.transaction_date = datetime.strptime(transaction_date, '%Y-%m-%dT%H:%M')
        transaction.counterparty = counterparty
        transaction.description = description if description else counterparty
        
        # 거래 유형에 따른 처리
        amount_value = float(amount)
        if transaction_type == 'deposit':
            transaction.transaction_type = 'credit'
            transaction.amount = amount_value
        elif transaction_type == 'withdrawal':
            transaction.transaction_type = 'debit'
            transaction.amount = -amount_value  # 출금은 음수
        elif transaction_type == 'transfer':
            transaction.transaction_type = 'transfer'
            transaction.amount = -amount_value  # 이체도 음수 (보내는 쪽)
            if target_account_id:
                target_account = Account.query.get(target_account_id)
                if target_account:
                    transaction.description += f' (받는 계정: {target_account.name})'
                    # 받는 계정 ID를 별도 필드에 저장 (향후 확장용)
                    transaction.counterpart_account = target_account.account_number if target_account.account_number else target_account.name
        
        # 분류 정보 설정 (이체가 아닌 경우만)
        if transaction_type != 'transfer':
            transaction.category_id = category_id if category_id else None
            transaction.department_id = department_id if department_id else None
            transaction.vendor_id = vendor_id if vendor_id else None
            transaction.classification_status = 'manual'  # 수동 추가
        else:
            transaction.classification_status = 'classified'  # 이체는 자동 분류됨
        
        db.session.add(transaction)
        db.session.commit()
        
        # 이상거래 알림 확인 및 생성
        check_and_create_anomaly_alerts(transaction)
        
        flash('거래가 성공적으로 추가되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'거래 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('transactions'))

@app.route('/download-sample/<format>')
@login_required
def download_sample(format):
    """샘플 파일 다운로드"""
    try:
        # 샘플 데이터 생성
        sample_data = {
            '계정': ['KB국민은행 주거래', 'KB국민은행 주거래', '신한카드 법인카드', 'KB국민은행 주거래', '우리은행 급여계좌'],  # 필수
            '거래일자': ['2024-01-15', '2024-01-16', '2024-01-17', '2024-01-18', '2024-01-19'],  # 필수
            '거래유형': ['출금', '입금', '출금', '이체', '입금'],  # 필수
            '금액': [5500, 150000, 12000, 50000, 200000],  # 필수
            '거래처': ['스타벅스 강남점', '클라이언트 A', '오피스디포', '내 저축계좌', '회사 급여'],  # 필수
            '거래시간': ['09:30:00', '14:15:30', '16:45:20', '11:20:00', '13:30:00'],  # 선택사항
            '메모': ['커피 및 간식', '프로젝트 수수료', '사무용품 구매', '저축 이체', '월급'],  # 선택사항
            '대상계정': ['', '', '', 'KB국민은행 저축계좌', ''],  # 이체 시 필수
            '분류': ['식음료', '매출', '사무용품', '', '급여'],  # 선택사항
            '부서': ['마케팅', '영업', '총무', '', '인사'],  # 선택사항
            '업체': ['스타벅스', '클라이언트A', '오피스디포', '', '회사']  # 선택사항
        }
        
        df = pd.DataFrame(sample_data)
        
        # 임시 파일 생성
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{format}') as tmp_file:
            if format == 'csv':
                df.to_csv(tmp_file.name, index=False, encoding='utf-8-sig')
                filename = 'bank_transactions_sample.csv'
                mimetype = 'text/csv'
            elif format == 'excel':
                df.to_excel(tmp_file.name, index=False, engine='openpyxl')
                filename = 'bank_transactions_sample.xlsx'
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                return jsonify({'error': '지원하지 않는 형식입니다.'}), 400
            
            return send_file(tmp_file.name, 
                           as_attachment=True, 
                           download_name=filename,
                           mimetype=mimetype)
                           
    except Exception as e:
        return jsonify({'error': f'샘플 파일 생성 중 오류가 발생했습니다: {str(e)}'}), 500


@app.route('/verify-uploads', methods=['POST'])
@login_required
def verify_uploads():
    """전체 업로드 검증"""
    try:
        # 총 거래 건수 조회
        total_transactions = Transaction.query.count()
        
        # 최근 업로드들의 상태 확인 (최근 7일)
        from datetime import datetime, timedelta
        recent_count = Transaction.query.filter(
            Transaction.created_at >= datetime.now() - timedelta(days=7)
        ).count()
        
        return jsonify({
            'success': True,
            'total_transactions': total_transactions,
            'recent_transactions': recent_count,
            'message': f'최근 7일간 {recent_count}건의 거래가 업로드되었습니다.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/check-upload/<int:upload_id>', methods=['GET'])
@login_required
def check_upload(upload_id):
    """특정 업로드 확인"""
    try:
        # Transaction ID를 기준으로 해당 거래 조회
        transaction = Transaction.query.get(upload_id)
        
        if not transaction:
            return jsonify({'success': False, 'error': '업로드를 찾을 수 없습니다.'})
        
        # 같은 날짜, 같은 계정의 거래 건수 조회
        same_date_count = Transaction.query.filter(
            func.date(Transaction.created_at) == func.date(transaction.created_at),
            Transaction.account_id == transaction.account_id
        ).count()
        
        return jsonify({
            'success': True,
            'filename': f"transactions_{transaction.created_at.strftime('%Y-%m-%d')}.csv",
            'processed_count': same_date_count,
            'upload_time': transaction.created_at.strftime('%Y-%m-%d %H:%M'),
            'account_name': transaction.account.name if transaction.account else 'N/A'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete-upload/<int:upload_id>', methods=['POST'])
@login_required
def delete_upload(upload_id):
    """업로드 기록 삭제"""
    try:
        # 해당 업로드 날짜의 모든 거래를 찾아서 삭제
        target_transaction = Transaction.query.get(upload_id)
        
        if not target_transaction:
            return jsonify({
                'success': False,
                'error': '해당 업로드 기록을 찾을 수 없습니다.'
            })
        
        # 같은 날짜, 같은 계정의 모든 거래를 삭제
        upload_date = target_transaction.created_at.date()
        account_id = target_transaction.account_id
        
        # 해당 날짜에 생성된 모든 거래 삭제
        transactions_to_delete = Transaction.query.filter(
            func.date(Transaction.created_at) == upload_date,
            Transaction.account_id == account_id
        ).all()
        
        deleted_count = len(transactions_to_delete)
        
        # 거래들을 소프트 삭제 (is_active = False)
        for transaction in transactions_to_delete:
            transaction.is_active = False
        
        db.session.commit()
        
        # 감사 로그 추가
        audit_log = AuditLog()
        audit_log.user_id = current_user.id
        audit_log.action = 'delete_upload'
        audit_log.table_name = 'transactions'
        audit_log.record_id = upload_id
        audit_log.new_values = f'업로드 기록 삭제: {deleted_count}건의 거래 삭제'
        audit_log.created_at = datetime.now()
        
        db.session.add(audit_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count}건의 거래가 성공적으로 삭제되었습니다.',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'삭제 중 오류가 발생했습니다: {str(e)}'
        })


def create_tables():
    """앱 시작시 테이블 생성 및 초기 데이터 로드"""
    db.create_all()
    init_sample_data()
